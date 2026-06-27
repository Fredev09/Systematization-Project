"""
Vistas paralelas para Productos usando DynamicService.

Estas vistas reemplazan internamente Producto.objects por DynamicService,
pero mantienen las mismas URLs, parametros GET, variables de template
y experiencia de usuario que las vistas legacy en views.py.

Uso temporal: se agregan como rutas alternativas en config/urls.py
para validacion antes de eliminar las vistas legacy.
"""

import logging
from decimal import Decimal

logger = logging.getLogger(__name__)

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from config.pagination import OPCIONES_POR_PAGINA, obtener_por_pagina, parametros_sin_pagina
from config.permissions import admin_required, es_administrador, rol_usuario
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from apps.platform.dynamic_forms.models import Campo, Formulario, Registro, ValorCampo
from apps.platform.dynamic_forms.services_dynamic import (
    DynamicService as DS,
    ValidacionError,
    FORM_MOVIMIENTOS_INVENTARIO,
    FORM_PRODUCTOS,
)
from .wrappers import DynamicMovimientoInventarioWrapper, DynamicProductWrapper


# ======================================================================
# HELPERS
# ======================================================================



def _stock_stats():
    """
    Calcula estadisticas de stock usando DynamicService.
    Retorna (total_productos, stock_bajo, sin_stock).
    """
    total = DS.contar(FORM_PRODUCTOS)
    # Stock bajo: registros con stock entre 1 y 5
    # Sin stock: registros con stock = 0
    stock_bajo = 0
    sin_stock = 0

    registros = Registro.objects.filter(
        formulario=DS.obtener_formulario(FORM_PRODUCTOS)
    )
    valores_map = DS.cargar_valores_mapa(registros)

    for r in registros:
        vals = valores_map.get(r.id, {})
        stock_str = vals.get('stock', '0')
        try:
            stock = int(stock_str)
        except (ValueError, TypeError):
            stock = 0
        if stock == 0:
            sin_stock += 1
        elif stock <= 5:
            stock_bajo += 1

    return total, stock_bajo, sin_stock


# ======================================================================
# VISTAS
# ======================================================================


@login_required(login_url='login')
@admin_required
def listar_productos(request):
    """
    Listado de productos usando DynamicService.
    PARAMETROS GET (compatibles con la vista legacy):
        q: texto de busqueda
        (cualquier campo activo del formulario se puede filtrar por GET)
        stock: 'bajo', 'sin_stock', o vacio
        page: numero de pagina
        per_page: items por pagina
    """
    query = request.GET.get('q', '').strip()
    stock_filtro = request.GET.get('stock', '').strip()

    # Campos activos del formulario Productos (para filtros dinámicos y templates)
    campos_producto = DS.obtener_campos_activos(FORM_PRODUCTOS)

    # Obtener registros del formulario Productos
    registros = Registro.objects.filter(
        formulario=DS.obtener_formulario(FORM_PRODUCTOS)
    ).order_by('-fecha_creacion')

    # Precargar todos los valores para filtrado y wrappeo
    todos_valores = DS.cargar_valores_mapa(registros)

    # Leer filtros dinámicos desde GET params para todos los campos del formulario
    filtros_activos = {}
    for campo in campos_producto:
        valor = request.GET.get(campo.nombre, '').strip()
        if valor:
            filtros_activos[campo.nombre] = valor

    # 1. Filtro por búsqueda textual (q) sobre todos los campos de tipo texto
    if query:
        query_lower = query.lower()
        campos_texto = [c for c in campos_producto if c.tipo in ('texto', 'email', 'url', 'telefono', 'textarea')]
        ids_filtrados = set()
        for r in registros:
            vals = todos_valores.get(r.id, {})
            for ct in campos_texto:
                if query_lower in vals.get(ct.nombre, '').lower():
                    ids_filtrados.add(r.id)
                    break
        registros = [r for r in registros if r.id in ids_filtrados]
    else:
        registros = list(registros)

    # 2. Filtros dinámicos por campo (lista, booleano, texto exacto, etc.)
    for campo_nombre, valor_buscado in filtros_activos.items():
        registros = [
            r for r in registros
            if todos_valores.get(r.id, {}).get(campo_nombre, '') == valor_buscado
        ]

    # 3. Filtro por stock (no es un campo del producto, es filtro de inventario)
    if stock_filtro == 'bajo':
        registros = [
            r for r in registros
            if 1 <= _entero(todos_valores.get(r.id, {}).get('stock', '0')) <= 5
        ]
    elif stock_filtro == 'sin_stock':
        registros = [
            r for r in registros
            if _entero(todos_valores.get(r.id, {}).get('stock', '0')) == 0
        ]

    # Envolver en wrappers para compatibilidad con templates
    productos = [
        DynamicProductWrapper(r, todos_valores.get(r.id, {}))
        for r in registros
    ]

    # Estadisticas
    total_productos = DS.contar(FORM_PRODUCTOS)
    _, stock_bajo, sin_stock = _stock_stats()

    # Paginacion
    per_page, per_page_int = obtener_por_pagina(request)
    paginator = Paginator(productos, per_page_int)
    pagina = request.GET.get('page')
    productos_pagina = paginator.get_page(pagina)
    query_params = parametros_sin_pagina(request, ['page'])

    # Categorias para el filtro (extraídas dinámicamente de campos_producto)
    from types import SimpleNamespace
    categorias = []
    for campo in campos_producto:
        if campo.nombre == 'categoria' and campo.opciones:
            categorias = [
                SimpleNamespace(id=op, nombre=op)
                for op in campo.opciones
            ]
            break
    categoria_filtro = filtros_activos.get('categoria', '')

    return render(request, 'productos/productos.html', {
        'productos': productos_pagina,
        'categorias': categorias,
        'query': query,
        'categoria_id': categoria_filtro,
        'stock': stock_filtro,
        'total_productos': total_productos,
        'stock_bajo': stock_bajo,
        'sin_stock': sin_stock,
        'stock_minimo_alerta': 5,
        'query_params': query_params,
        'per_page': per_page,
        'per_page_options': OPCIONES_POR_PAGINA,
        'campos_producto': campos_producto,
        'es_admin': es_administrador(request.user),
        'rol_usuario': rol_usuario(request.user),
    })


# ======================================================================
# HELPER: ESTADÍSTICAS COMPLETAS DE INVENTARIO
# ======================================================================


def _stock_stats_completo(stock_minimo=5):
    """
    Calcula todas las estadísticas de inventario usando DynamicService.

    A diferencia de _stock_stats(), este helper:
    - Acepta stock_minimo como parámetro (desde ConfiguracionTienda)
    - Incluye disponibles y valor_total
    - Retorna un dict nombrado para facilitar el contexto de templates

    Args:
        stock_minimo: Umbral para considerar stock bajo (default: 5).

    Returns:
        dict con: total_productos, stock_bajo, sin_stock,
                 disponibles, valor_total
    """
    total_productos = DS.contar(FORM_PRODUCTOS)

    registros = Registro.objects.filter(
        formulario=DS.obtener_formulario(FORM_PRODUCTOS)
    )
    valores_map = DS.cargar_valores_mapa(registros)

    stock_bajo = 0
    sin_stock = 0
    disponibles = 0
    valor_total = Decimal('0')

    for r in registros:
        vals = valores_map.get(r.id, {})

        stock_str = vals.get('stock', '0')
        try:
            stock = int(stock_str)
        except (ValueError, TypeError):
            stock = 0

        precio_str = vals.get('precio', '0')
        try:
            precio = Decimal(str(precio_str).replace(',', '.'))
        except (ValueError, TypeError):
            precio = Decimal('0')

        valor_total += precio * stock

        if stock == 0:
            sin_stock += 1
        elif stock <= stock_minimo:
            stock_bajo += 1
        else:
            disponibles += 1

    return {
        'total_productos': total_productos,
        'stock_bajo': stock_bajo,
        'sin_stock': sin_stock,
        'disponibles': disponibles,
        'valor_total': valor_total,
    }


# ======================================================================
# HELPER: MOVIMIENTOS RECIENTES DE INVENTARIO
# ======================================================================


def _movimientos_recientes_dinamicos(limite=50, producto_wrappers=None):
    """
    Obtiene los movimientos de inventario recientes desde el formulario
    dinámico MovimientosInventario, envueltos en DynamicMovimientoInventarioWrapper.

    Sigue el mismo patrón que _envolver_ventas() en ventas/views_dynamic.py:
    - Una consulta para los registros de movimientos
    - Una consulta batch para cargar valores
    - Una consulta batch para resolver productos relacionados

    Args:
        limite: Cantidad máxima de movimientos a retornar (default: 50).
                Pasar None para obtener todos los movimientos sin límite.
        producto_wrappers: Opcional, dict {registro_id: DynamicProductWrapper}
                           pre-resuelto. Si no se pasa, se resuelve automáticamente.

    Returns:
        Lista de DynamicMovimientoInventarioWrapper.
    """
    try:
        form = DS.obtener_formulario(FORM_MOVIMIENTOS_INVENTARIO)
    except Exception:
        return []

    qs = Registro.objects.filter(
        formulario=form
    ).order_by('-fecha_creacion')
    if limite is not None:
        qs = qs[:limite]
    registros = qs

    if not registros:
        return []

    # Cargar valores de todos los movimientos
    valores_map = DS.cargar_valores_mapa(registros)

    # Resolver productos relacionados si no se pasaron pre-resueltos
    if producto_wrappers is None:
        producto_ids = set()
        for r in registros:
            vals = valores_map.get(r.id, {})
            prod_id = vals.get('producto', '').strip()
            if prod_id and prod_id.isdigit():
                producto_ids.add(int(prod_id))

        producto_wrappers = {}
        if producto_ids:
            form_productos = DS.obtener_formulario(FORM_PRODUCTOS)
            prod_registros = Registro.objects.filter(
                id__in=list(producto_ids),
                formulario=form_productos
            )
            prod_valores = DS.cargar_valores_mapa(prod_registros)
            for pr in prod_registros:
                producto_wrappers[pr.id] = DynamicProductWrapper(
                    pr, prod_valores.get(pr.id, {})
                )

    movimientos = []
    for r in registros:
        vals = valores_map.get(r.id, {})
        prod_id_str = vals.get('producto', '').strip()
        prod_id = int(prod_id_str) if prod_id_str and prod_id_str.isdigit() else None
        pw = producto_wrappers.get(prod_id) if prod_id else None

        movimientos.append(DynamicMovimientoInventarioWrapper(
            r, vals, producto_wrapper=pw
        ))

    return movimientos


# ======================================================================
# HELPER: PRODUCTOS CRÍTICOS (STOCK BAJO / SIN STOCK)
# ======================================================================


def _productos_criticos_dinamicos(stock_minimo=5, limite=8):
    """
    Obtiene los productos cuyo stock está entre 0 y stock_minimo (inclusive),
    ordenados por stock ascendente y luego por nombre.

    Reutiliza DynamicService para carga en lote y DynamicProductWrapper
    para compatibilidad con templates legacy.

    Args:
        stock_minimo: Umbral máximo de stock para considerar crítico (default: 5).
        limite: Cantidad máxima de productos a retornar (default: 8).

    Returns:
        Lista de DynamicProductWrapper ordenada por stock ascendente.
    """
    registros = Registro.objects.filter(
        formulario=DS.obtener_formulario(FORM_PRODUCTOS)
    )
    valores_map = DS.cargar_valores_mapa(registros)

    # Filtrar productos con stock entre 0 y stock_minimo
    criticos = []
    for r in registros:
        vals = valores_map.get(r.id, {})
        stock_str = vals.get('stock', '0')
        try:
            stock = int(stock_str)
        except (ValueError, TypeError):
            stock = 0

        if 0 <= stock <= stock_minimo:
            pw = DynamicProductWrapper(r, vals)
            criticos.append(pw)

    # Ordenar por stock ascendente, luego por nombre
    criticos.sort(key=lambda p: (p.stock, p.nombre.lower()))

    return criticos[:limite]


# ======================================================================
# VISTA: INVENTARIO (DINÁMICO)
# ======================================================================


@login_required(login_url='login')
@admin_required
def inventario(request):
    """
    Vista de inventario usando exclusivamente Dynamic Forms.

    Reutiliza los 3 helpers del módulo:
    - _stock_stats_completo() para estadísticas
    - _productos_criticos_dinamicos() para productos con prioridad
    - _movimientos_recientes_dinamicos() para movimientos recientes

    Renderiza el mismo template inventario/inventario.html con
    contexto compatible.
    """
    # ------------------------------------------------------------------
    # Parámetros y configuración
    # ------------------------------------------------------------------
    query = request.GET.get('q', '').strip()
    stock_filtro = request.GET.get('stock', '').strip()

    try:
        from apps.shared.configuracion.models import ConfiguracionTienda
        stock_minimo = ConfiguracionTienda.obtener().stock_minimo_alerta
    except Exception:
        stock_minimo = 5

    # Campos activos del formulario Productos (para filtros dinámicos y templates)
    campos_producto = DS.obtener_campos_activos(FORM_PRODUCTOS)

    # ------------------------------------------------------------------
    # Estadísticas completas
    # ------------------------------------------------------------------
    stats = _stock_stats_completo(stock_minimo)

    # ------------------------------------------------------------------
    # Productos críticos (stock <= stock_minimo)
    # ------------------------------------------------------------------
    productos_criticos = _productos_criticos_dinamicos(stock_minimo, limite=8)

    # ------------------------------------------------------------------
    # Listado de productos con filtros dinámicos
    # ------------------------------------------------------------------
    registros = Registro.objects.filter(
        formulario=DS.obtener_formulario(FORM_PRODUCTOS)
    )

    todos_valores = DS.cargar_valores_mapa(registros)

    # Leer filtros dinámicos desde GET params para todos los campos del formulario
    filtros_activos = {}  # {nombre_campo: valor_buscado}
    for campo in campos_producto:
        valor = request.GET.get(campo.nombre, '').strip()
        if valor:
            filtros_activos[campo.nombre] = valor

    # 1. Filtro por búsqueda textual (q) sobre todos los campos de tipo texto
    if query:
        query_lower = query.lower()
        campos_texto = [c for c in campos_producto if c.tipo in ('texto', 'email', 'url', 'telefono', 'textarea')]
        ids_filtrados = set()
        for r in registros:
            vals = todos_valores.get(r.id, {})
            for ct in campos_texto:
                if query_lower in vals.get(ct.nombre, '').lower():
                    ids_filtrados.add(r.id)
                    break
        registros = [r for r in registros if r.id in ids_filtrados]
    else:
        registros = list(registros)

    # 2. Filtros dinámicos por campo (lista, booleano, texto exacto, etc.)
    for campo_nombre, valor_buscado in filtros_activos.items():
        registros = [
            r for r in registros
            if todos_valores.get(r.id, {}).get(campo_nombre, '') == valor_buscado
        ]

    # 3. Filtro por stock (no es un campo del producto, es filtro de inventario)
    if stock_filtro == 'bajo':
        registros = [
            r for r in registros
            if 1 <= _entero(todos_valores.get(r.id, {}).get('stock', '0')) <= stock_minimo
        ]
    elif stock_filtro == 'sin_stock':
        registros = [
            r for r in registros
            if _entero(todos_valores.get(r.id, {}).get('stock', '0')) == 0
        ]
    elif stock_filtro == 'disponible':
        registros = [
            r for r in registros
            if _entero(todos_valores.get(r.id, {}).get('stock', '0')) > stock_minimo
        ]

    # Envolver en wrappers
    productos = [
        DynamicProductWrapper(r, todos_valores.get(r.id, {}))
        for r in registros
    ]
    # Ordenar en Python (no se puede order_by campos EAV a nivel DB)
    productos.sort(key=lambda p: (p.stock, p.nombre.lower()))

    # Paginación de productos
    productos_per_page, productos_per_page_int = obtener_por_pagina(request, 'productos_per_page')
    productos_paginator = Paginator(productos, productos_per_page_int)
    productos_pagina_numero = request.GET.get('productos_page')
    productos_pagina = productos_paginator.get_page(productos_pagina_numero)
    productos_query_params = parametros_sin_pagina(request, ['productos_page'])

    # ------------------------------------------------------------------
    # Movimientos recientes (reutiliza _movimientos_recientes_dinamicos)
    # ------------------------------------------------------------------
    try:
        movimientos_todos = _movimientos_recientes_dinamicos(limite=None)
        total_movs = len(movimientos_todos)
        total_entradas = sum(1 for m in movimientos_todos if m.tipo == 'entrada')
        total_salidas = sum(1 for m in movimientos_todos if m.tipo == 'salida')
        total_correcciones = sum(1 for m in movimientos_todos if m.tipo == 'correccion')

        # Paginación de movimientos
        movimientos_per_page, movimientos_per_page_int = obtener_por_pagina(request, 'movimientos_per_page')
        mov_paginator = Paginator(movimientos_todos, movimientos_per_page_int)
        mov_pagina_numero = request.GET.get('movimientos_page')
        movimientos_recientes = mov_paginator.get_page(mov_pagina_numero)
        movimientos_query_params = parametros_sin_pagina(request, ['movimientos_page'])

    except Exception:
        movimientos_recientes = []
        total_movs = 0
        total_entradas = 0
        total_salidas = 0
        total_correcciones = 0
        movimientos_per_page = 10
        movimientos_query_params = ''

    # ------------------------------------------------------------------
    # Render
    # ------------------------------------------------------------------
    return render(request, 'inventario/inventario.html', {
        'productos': productos_pagina,
        'query': query,
        'stock': stock_filtro,
        'total_productos': stats['total_productos'],
        'stock_bajo': stats['stock_bajo'],
        'sin_stock': stats['sin_stock'],
        'disponibles': stats['disponibles'],
        'valor_total': stats['valor_total'],
        'productos_criticos': productos_criticos,
        'stock_minimo_alerta': stock_minimo,
        'movimientos_recientes': movimientos_recientes,
        'total_movimientos': total_movs,
        'total_entradas': total_entradas,
        'total_salidas': total_salidas,
        'total_correcciones': total_correcciones,
        'productos_query_params': productos_query_params,
        'movimientos_query_params': movimientos_query_params,
        'productos_per_page': productos_per_page,
        'movimientos_per_page': movimientos_per_page,
        'per_page_options': OPCIONES_POR_PAGINA,
        'campos_producto': campos_producto,
        'es_admin': es_administrador(request.user),
        'rol_usuario': rol_usuario(request.user),
    })


# ======================================================================
# VISTA: ACTUALIZAR STOCK (DINÁMICO)
# ======================================================================

# Mapa de motivos JS → valores del formulario dinámico MovimientosInventario.
# El JS en actualizar_stock.js envía valores como 'compra_proveedor' pero
# el formulario dinámico almacena el texto display completo.
_MOTIVO_DINAMICO_MAP = {
    'compra_proveedor': 'Compra a proveedor',
    'devolucion_cliente': 'Devolucion de cliente',
    'venta_manual': 'Venta manual',
    'venta_sistema': 'Venta del sistema',
    'producto_danado': 'Producto danado',
    'perdida': 'Perdida',
    'robo': 'Robo',
    'devolucion_proveedor': 'Devolucion a proveedor',
    'conteo_fisico': 'Conteo fisico',
    'correccion_manual': 'Correccion manual',
}

_TIPO_DINAMICO_MAP = {
    'entrada': 'Entrada',
    'salida': 'Salida',
    'correccion': 'Correccion',
}


@login_required(login_url='login')
@admin_required
def actualizar_stock(request, producto_id):
    """
    Ajusta el stock de un producto usando Dynamic Forms.

    Reimplementación dinámica de actualizar_stock legacy.
    Usa select_for_update + transaction.atomic() para evitar
    condiciones de carrera en la actualización concurrente del stock.

    Args:
        producto_id: ID del Registro del formulario Productos.
    """
    formulario = DS.obtener_formulario(FORM_PRODUCTOS)
    registro = get_object_or_404(Registro, id=producto_id, formulario=formulario)
    valores = DS.obtener_valores(registro)
    producto = DynamicProductWrapper(registro, valores)

    try:
        from apps.shared.configuracion.models import ConfiguracionTienda
        stock_minimo = ConfiguracionTienda.obtener().stock_minimo_alerta
    except Exception:
        stock_minimo = 5

    if request.method == 'POST':
        tipo = request.POST.get('tipo', '').strip()
        motivo = request.POST.get('motivo', '').strip() or None
        cantidad_raw = request.POST.get('cantidad', '').strip()
        observacion = request.POST.get('observacion', '').strip() or None

        # --- Validaciones (idénticas al legacy) ---
        if not tipo or not cantidad_raw:
            messages.error(request, 'Debes completar tipo de movimiento y cantidad.')
            return redirect('actualizar_stock', producto_id=producto_id)

        try:
            cantidad = int(cantidad_raw)
        except ValueError:
            messages.error(request, 'La cantidad debe ser un numero valido.')
            return redirect('actualizar_stock', producto_id=producto_id)

        if cantidad < 0:
            messages.error(request, 'La cantidad no puede ser negativa.')
            return redirect('actualizar_stock', producto_id=producto_id)

        if tipo in ('entrada', 'salida') and cantidad == 0:
            messages.error(request, 'Para entradas o salidas la cantidad debe ser mayor a 0.')
            return redirect('actualizar_stock', producto_id=producto_id)

        if tipo not in ('entrada', 'salida', 'correccion'):
            messages.error(request, 'El tipo de movimiento no es valido.')
            return redirect('actualizar_stock', producto_id=producto_id)

        # Mapear valores JS → valores del formulario dinámico
        tipo_dinamico = _TIPO_DINAMICO_MAP[tipo]
        motivo_dinamico = _MOTIVO_DINAMICO_MAP.get(motivo) if motivo else None

        try:
            with transaction.atomic():
                # 1. Bloquear el registro del producto para escritura concurrente
                producto_bloqueado = Registro.objects.select_for_update().get(id=producto_id)

                # 2. Leer stock actual desde ValorCampo
                stock_anterior = _entero(
                    DS.obtener_valor(producto_bloqueado, 'stock', '0')
                )

                # 3. Calcular nuevo stock
                if tipo == 'entrada':
                    stock_nuevo = stock_anterior + cantidad
                    cantidad_movimiento = cantidad
                elif tipo == 'salida':
                    stock_nuevo = stock_anterior - cantidad
                    cantidad_movimiento = cantidad
                    if stock_nuevo < 0:
                        raise ValidationError(
                            'No puedes retirar mas productos de los que hay actualmente.'
                        )
                else:  # correccion
                    stock_nuevo = cantidad
                    cantidad_movimiento = abs(stock_nuevo - stock_anterior)

                # 4. Actualizar el stock vía DS.actualizar (sin lock propio)
                DS.actualizar(
                    producto_bloqueado,
                    {'stock': str(stock_nuevo)},
                    usuario=request.user,
                    usar_select_for_update=False,
                )

                # 5. Crear el movimiento de inventario
                DS.crear(
                    FORM_MOVIMIENTOS_INVENTARIO,
                    {
                        'producto': str(producto_id),
                        'tipo': tipo_dinamico,
                        'cantidad': str(cantidad_movimiento),
                        'motivo': motivo_dinamico or '',
                        'stock_anterior': str(stock_anterior),
                        'stock_nuevo': str(stock_nuevo),
                        'observacion': observacion or '',
                    },
                    usuario=request.user,
                )

            messages.success(
                request,
                f'Stock de "{producto.nombre}" actualizado correctamente. '
                f'Antes: {stock_anterior}, ahora: {stock_nuevo}.'
            )
            return redirect('inventario')

        except ValidacionError as e:
            for error in e.errores:
                messages.error(request, error)
            return redirect('actualizar_stock', producto_id=producto_id)

        except ValidationError as e:
            messages.error(request, e.messages[0] if e.messages else str(e))
            return redirect('actualizar_stock', producto_id=producto_id)

        except Exception as e:
            logger.exception(f'Error actualizando stock del producto #{producto_id}: {e}')
            messages.error(request, f'Error inesperado: {e}')
            return redirect('actualizar_stock', producto_id=producto_id)

    # ==================================================================
    # GET: Renderizar el formulario
    # ==================================================================

    # Cargar movimientos de este producto en el formulario dinámico
    try:
        form_mov = DS.obtener_formulario(FORM_MOVIMIENTOS_INVENTARIO)
        campo_producto = DS.obtener_campo(form_mov, 'producto')
        mov_registros = Registro.objects.filter(
            formulario=form_mov,
            valores__campo=campo_producto,
            valores__valor=str(producto_id),
        ).order_by('-fecha_creacion')

        mov_valores = DS.cargar_valores_mapa(mov_registros)

        per_page, per_page_int = obtener_por_pagina(request)
        mov_paginator = Paginator(list(mov_registros), per_page_int)
        pagina = request.GET.get('page')
        mov_pagina = mov_paginator.get_page(pagina)

        movimientos_producto = []
        for mr in mov_pagina:
            vals = mov_valores.get(mr.id, {})
            movimientos_producto.append(
                DynamicMovimientoInventarioWrapper(mr, vals)
            )

        query_params = parametros_sin_pagina(request, ['page'])

    except Exception:
        movimientos_producto = []
        per_page = 10
        query_params = ''

    # Campos activos del formulario Productos (para templates dinámicos)
    campos_producto = DS.obtener_campos_activos(FORM_PRODUCTOS)

    # Opciones de tipo compatibles con el template legacy
    # El template espera (valor_lowercase, texto_display)
    tipoElecciones = [
        ('entrada', 'Entrada'),
        ('salida', 'Salida'),
        ('correccion', 'Correccion de stock'),
    ]

    return render(request, 'productos/actualizar_stock.html', {
        'producto': producto,
        'tipoElecciones': tipoElecciones,
        'movimientos_producto': movimientos_producto,
        'stock_minimo_alerta': stock_minimo,
        'campos_producto': campos_producto,
        'query_params': query_params,
        'per_page': per_page,
        'per_page_options': OPCIONES_POR_PAGINA,
        'es_admin': es_administrador(request.user),
        'rol_usuario': rol_usuario(request.user),
    })


# ======================================================================
# HELPER COMPARTIDO: FILTRAR MOVIMIENTOS (usado por historial y exportación)
# ======================================================================


def _filtrar_movimientos_dinamicos(request):
    """
    Filtra movimientos de inventario según parámetros GET y retorna
    la lista envuelta lista para templates/exportación.

    Extraído como helper compartido para evitar duplicación entre
    historial_inventario y exportar_historial_inventario_excel.

    Args:
        request: HttpRequest con posibles GET params: q, tipo, motivo,
                 fecha_inicio, fecha_fin

    Returns:
        (mov_list, filtros) donde:
            mov_list: Lista de DynamicMovimientoInventarioWrapper
            filtros: Dict con query, tipo, motivo, fecha_inicio, fecha_fin,
                     total_filtrado, entradas_filtradas, salidas_filtradas,
                     correcciones_filtradas
        Si no hay resultados o no existen formularios, retorna ([], filtros_vacios)
    """
    query = request.GET.get('q', '').strip()
    tipo = request.GET.get('tipo', '').strip()
    motivo = request.GET.get('motivo', '').strip()
    fecha_inicio = request.GET.get('fecha_inicio', '').strip()
    fecha_fin = request.GET.get('fecha_fin', '').strip()

    filtros_base = {
        'query': query,
        'tipo': tipo,
        'motivo': motivo,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
    }

    try:
        form_mov = DS.obtener_formulario(FORM_MOVIMIENTOS_INVENTARIO)
        form_prod = DS.obtener_formulario(FORM_PRODUCTOS)
    except Exception:
        return [], {**filtros_base, 'total_filtrado': 0,
                     'entradas_filtradas': 0, 'salidas_filtradas': 0,
                     'correcciones_filtradas': 0}

    mov_registros = Registro.objects.filter(formulario=form_mov)

    if tipo:
        tipo_dinamico = _TIPO_DINAMICO_MAP.get(tipo)
        if tipo_dinamico:
            campo_tipo = DS.obtener_campo(form_mov, 'tipo')
            mov_registros = mov_registros.filter(
                valores__campo=campo_tipo, valores__valor=tipo_dinamico
            )

    if motivo:
        motivo_dinamico = _MOTIVO_DINAMICO_MAP.get(motivo)
        if motivo_dinamico:
            campo_motivo = DS.obtener_campo(form_mov, 'motivo')
            mov_registros = mov_registros.filter(
                valores__campo=campo_motivo, valores__valor=motivo_dinamico
            )

    if fecha_inicio:
        mov_registros = mov_registros.filter(fecha_creacion__date__gte=fecha_inicio)
    if fecha_fin:
        mov_registros = mov_registros.filter(fecha_creacion__date__lte=fecha_fin)

    if query:
        campos_producto = DS.obtener_campos_activos(FORM_PRODUCTOS)
        campos_texto = [c for c in campos_producto if c.tipo in ('texto', 'email', 'url', 'telefono', 'textarea')]
        prod_registros = Registro.objects.filter(formulario=form_prod)
        prod_valores = DS.cargar_valores_mapa(prod_registros)
        matching_prod_ids = set()
        for r in prod_registros:
            vals = prod_valores.get(r.id, {})
            for ct in campos_texto:
                if query.lower() in vals.get(ct.nombre, '').lower():
                    matching_prod_ids.add(str(r.id))
                    break

        campo_producto = DS.obtener_campo(form_mov, 'producto')
        campo_observacion = DS.obtener_campo(form_mov, 'observacion')
        q_parts = Q(valores__campo=campo_observacion, valores__valor__icontains=query)
        if matching_prod_ids:
            q_parts |= Q(valores__campo=campo_producto,
                         valores__valor__in=list(matching_prod_ids))
        mov_registros = mov_registros.filter(q_parts).distinct()

    mov_registros = mov_registros.order_by('-fecha_creacion')
    mov_ids = list(mov_registros.values_list('id', flat=True))

    if not mov_ids:
        return [], {**filtros_base, 'total_filtrado': 0,
                     'entradas_filtradas': 0, 'salidas_filtradas': 0,
                     'correcciones_filtradas': 0}

    mov_regs = Registro.objects.filter(id__in=mov_ids)
    mov_valores = DS.cargar_valores_mapa(mov_regs)

    entradas_filtradas = sum(1 for v in mov_valores.values()
                              if v.get('tipo', '') == 'Entrada')
    salidas_filtradas = sum(1 for v in mov_valores.values()
                             if v.get('tipo', '') == 'Salida')
    correcciones_filtradas = sum(1 for v in mov_valores.values()
                                  if v.get('tipo', '') == 'Correccion')

    producto_ids = set()
    for vals in mov_valores.values():
        pid = vals.get('producto', '').strip()
        if pid and pid.isdigit():
            producto_ids.add(int(pid))

    producto_wrappers = {}
    if producto_ids:
        prod_registros = Registro.objects.filter(
            id__in=list(producto_ids), formulario=form_prod
        )
        prod_valores = DS.cargar_valores_mapa(prod_registros)
        for pr in prod_registros:
            producto_wrappers[pr.id] = DynamicProductWrapper(
                pr, prod_valores.get(pr.id, {})
            )

    mov_list = []
    for r in mov_regs.order_by('-fecha_creacion'):
        vals = mov_valores.get(r.id, {})
        pid_str = vals.get('producto', '').strip()
        pid = int(pid_str) if pid_str and pid_str.isdigit() else None
        pw = producto_wrappers.get(pid) if pid else None
        mov_list.append(DynamicMovimientoInventarioWrapper(r, vals, producto_wrapper=pw))

    return mov_list, {**filtros_base,
                      'total_filtrado': len(mov_ids),
                      'entradas_filtradas': entradas_filtradas,
                      'salidas_filtradas': salidas_filtradas,
                      'correcciones_filtradas': correcciones_filtradas}


# ======================================================================
# VISTA: HISTORIAL DE INVENTARIO (DINÁMICO)
# ======================================================================


@login_required(login_url='login')
@admin_required
def historial_inventario(request):
    """
    Historial de movimientos de inventario con filtros.

    Reimplementación dinámica que reutiliza _filtrar_movimientos_dinamicos().
    """
    mov_list, filtros = _filtrar_movimientos_dinamicos(request)

    campos_producto = DS.obtener_campos_activos(FORM_PRODUCTOS)

    tipoElecciones = [
        ('entrada', 'Entrada'),
        ('salida', 'Salida'),
        ('correccion', 'Correccion de stock'),
    ]
    motivoElecciones = list(_MOTIVO_DINAMICO_MAP.items())

    if not mov_list:
        return render(request, 'inventario/historial_inventario.html', {
            'movimientos': [],
            **filtros,
            'tipoElecciones': tipoElecciones,
            'motivoElecciones': motivoElecciones,
            'campos_producto': campos_producto,
            'query_params': '',
            'per_page': '10',
            'per_page_options': OPCIONES_POR_PAGINA,
            'es_admin': es_administrador(request.user),
            'rol_usuario': rol_usuario(request.user),
        })

    per_page, per_page_int = obtener_por_pagina(request)
    paginator = Paginator(mov_list, per_page_int)
    pagina = request.GET.get('page')
    movimientos = paginator.get_page(pagina)
    query_params = parametros_sin_pagina(request, ['page'])

    return render(request, 'inventario/historial_inventario.html', {
        'movimientos': movimientos,
        **filtros,
        'tipoElecciones': tipoElecciones,
        'motivoElecciones': motivoElecciones,
        'campos_producto': campos_producto,
        'query_params': query_params,
        'per_page': per_page,
        'per_page_options': OPCIONES_POR_PAGINA,
        'es_admin': es_administrador(request.user),
        'rol_usuario': rol_usuario(request.user),
    })


# ======================================================================
# VISTA: EXPORTAR HISTORIAL A EXCEL (DINÁMICO)
# ======================================================================


@login_required(login_url='login')
@admin_required
def exportar_historial_inventario_excel(request):
    """
    Exporta el historial de inventario filtrado a Excel.

    Reimplementación dinámica de exportar_historial_inventario_excel legacy.
    Reutiliza _filtrar_movimientos_dinamicos() para obtener los mismos
    datos filtrados que la vista historial_inventario.

    Genera el mismo formato de archivo que el legacy:
    - Columnas: Producto, Categoria, Tipo, Cantidad, Stock anterior,
                Stock nuevo, Motivo, Observacion, Fecha
    - Título, subtítulo con filtros, encabezados con estilo
    - Congelamiento de paneles, filtro automático
    """
    mov_list, filtros = _filtrar_movimientos_dinamicos(request)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Historial inventario'

    # Encabezados de columna (orden = mismo que legacy)
    encabezados = [
        'Producto', 'Categoria', 'Tipo', 'Cantidad',
        'Stock anterior', 'Stock nuevo', 'Motivo',
        'Observacion', 'Fecha'
    ]

    # Fila 1: título
    ws.append(['Historial de movimientos de inventario'])

    # Fila 2: filtros aplicados
    ws.append([
        f"Busqueda: {filtros['query'] or 'Todos'}",
        f"Tipo: {filtros['tipo'] or 'Todos'}",
        f"Motivo: {filtros['motivo'] or 'Todos'}",
        f"Desde: {filtros['fecha_inicio'] or 'Sin fecha'}",
        f"Hasta: {filtros['fecha_fin'] or 'Sin fecha'}",
    ])

    # Fila 3: vacía
    ws.append([])

    # Fila 4: encabezados
    ws.append(encabezados)

    # Filas de datos
    for m in mov_list:
        nombre_prod = getattr(m.producto, 'nombre', '') or ''
        nombre_cat = getattr(m.producto, 'valores', {}).get('categoria', '')
        ws.append([
            nombre_prod,
            nombre_cat or 'Sin categoria',
            m.get_tipo_display(),
            m.cantidad,
            m.stock_anterior,
            m.stock_nuevo,
            m.get_motivo_display() if m.motivo else 'Sin motivo',
            m.observacion or '',
            timezone.localtime(m.fecha).strftime('%d/%m/%Y %I:%M %p') if m.fecha else '',
        ])

    # --- Estilos (idénticos al legacy) ---
    titulo_fill = PatternFill('solid', fgColor='D41473')
    encabezado_fill = PatternFill('solid', fgColor='FCE7F3')
    titulo_font = Font(color='FFFFFF', bold=True, size=14)
    encabezado_font = Font(color='111827', bold=True)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws['A1'].fill = titulo_fill
    ws['A1'].font = titulo_font
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 28

    for cell in ws[4]:
        cell.fill = encabezado_fill
        cell.font = encabezado_font
        cell.alignment = center

    for row in ws.iter_rows(min_row=5):
        for cell in row:
            cell.alignment = left

    anchos = {
        'A': 28, 'B': 22, 'C': 16, 'D': 12,
        'E': 16, 'F': 16, 'G': 26, 'H': 38, 'I': 22,
    }
    for columna, ancho in anchos.items():
        ws.column_dimensions[columna].width = ancho

    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:I{ws.max_row}'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="historial_inventario.xlsx"'
    wb.save(response)
    return response


# ======================================================================
# VISTA: AGREGAR PRODUCTO (DINÁMICO)
# ======================================================================


@login_required(login_url='login')
@admin_required
def agregar_producto(request):
    """
    Crea un producto usando DynamicService.

    Recolecta los valores del POST, pasa archivos via archivos_dict
    y delega toda la creación en DS.crear().
    """
    formulario = DS.obtener_formulario(FORM_PRODUCTOS)
    campos = formulario.campos.filter(activo=True).order_by('orden')

    errores = None
    valores_previos = {}

    if request.method == 'POST':
        # Construir valores_dict desde POST
        valores = {}
        archivos = {}
        for campo in campos:
            if campo.tipo in Campo.TIPOS_ARCHIVO:
                archivo = request.FILES.get(f'campo_{campo.id}')
                if archivo:
                    archivos[campo.nombre] = archivo
                # Para tipos archivo, también puede venir un valor POST
                # (ej: imagen_url se maneja por separado como campo url)
            elif campo.tipo not in ('calculado',):
                valor = request.POST.get(f'campo_{campo.id}', '').strip()
                if valor:
                    valores[campo.nombre] = valor
                valores_previos[campo.id] = valor

        try:
            registro = DS.crear(
                FORM_PRODUCTOS,
                valores_dict=valores,
                archivos_dict=archivos,
                usuario=request.user,
            )
            messages.success(
                request,
                f'Producto creado correctamente (ID #{registro.id}).'
            )
            return redirect('productos')

        except ValidacionError as e:
            errores = e.errores
            messages.error(request, 'Corrige los errores e intenta de nuevo.')
        except Exception as e:
            logger.exception(f'Error creando producto: {e}')
            errores = [f'Error inesperado: {e}']
            messages.error(request, 'No se pudo crear el producto.')

    return render(request, 'productos/agregar_producto_dinamico.html', {
        'formulario': formulario,
        'campos': campos,
        'errores': errores,
        'valores_previos': valores_previos,
        'es_admin': es_administrador(request.user),
        'rol_usuario': rol_usuario(request.user),
    })


# ======================================================================
# VISTA: EDITAR PRODUCTO (DINÁMICO)
# ======================================================================


@login_required(login_url='login')
@admin_required
def editar_producto(request, producto_id):
    """
    Edita un producto usando DynamicService.actualizar().

    Carga los valores actuales del registro, los pre-rellena en el formulario
    y delega la actualización en DS.actualizar() con soporte de archivos.

    Args:
        producto_id: ID del Registro del formulario Productos.
    """
    formulario = DS.obtener_formulario(FORM_PRODUCTOS)
    campos = formulario.campos.filter(activo=True).order_by('orden')

    registro = get_object_or_404(
        Registro.objects.select_related('formulario'),
        id=producto_id,
        formulario=formulario
    )

    # Cargar valores actuales del registro
    valores_actuales = DS.obtener_valores(registro)

    # Construir dict keyed por campo.id para el template
    valores_previos = {}
    imagenes_actuales = {}
    for campo in campos:
        valor = valores_actuales.get(campo.nombre, '')
        valores_previos[campo.id] = valor
        if campo.tipo in Campo.TIPOS_ARCHIVO and valor:
            imagenes_actuales[campo.id] = valor

    errores = None

    if request.method == 'POST':
        # Construir valores_dict desde POST
        valores = {}
        archivos = {}
        for campo in campos:
            if campo.tipo in Campo.TIPOS_ARCHIVO:
                archivo = request.FILES.get(f'campo_{campo.id}')
                if archivo:
                    archivos[campo.nombre] = archivo
            elif campo.tipo not in ('calculado',):
                valor = request.POST.get(f'campo_{campo.id}', '').strip()
                if valor:
                    valores[campo.nombre] = valor
                # Actualizar valores_previos con lo enviado
                valores_previos[campo.id] = valor

        try:
            registro = DS.actualizar(
                registro,
                valores_dict=valores,
                archivos_dict=archivos,
                usuario=request.user,
            )
            messages.success(
                request,
                f'Producto #{registro.id} actualizado correctamente.'
            )
            return redirect('productos')

        except ValidacionError as e:
            errores = e.errores
            messages.error(request, 'Corrige los errores e intenta de nuevo.')
        except Exception as e:
            logger.exception(f'Error actualizando producto #{producto_id}: {e}')
            errores = [f'Error inesperado: {e}']
            messages.error(request, 'No se pudo actualizar el producto.')

    return render(request, 'productos/editar_producto_dinamico.html', {
        'formulario': formulario,
        'campos': campos,
        'registro': registro,
        'errores': errores,
        'valores_previos': valores_previos,
        'imagenes_actuales': imagenes_actuales,
        'es_admin': es_administrador(request.user),
        'rol_usuario': rol_usuario(request.user),
    })


# ======================================================================
# VISTA: ELIMINAR PRODUCTO (DINÁMICO)
# ======================================================================


@login_required(login_url='login')
@admin_required
def eliminar_producto(request, producto_id):
    """
    Elimina un producto dinámico con verificación de integridad referencial.

    Antes de eliminar verifica:
        1. Que el producto exista como Registro.
        2. Que no tenga ventas asociadas en el formulario dinámico Ventas.
        3. Si hay ventas, muestra error y no permite eliminar.

    Usa DS.eliminar() que realiza el borrado físico del Registro
    (con cascade a ValorCampo).

    Args:
        producto_id: ID del Registro del formulario Productos.
    """
    formulario = DS.obtener_formulario(FORM_PRODUCTOS)
    registro = get_object_or_404(Registro, id=producto_id, formulario=formulario)

    # Cargar valores para el wrapper
    valores = DS.obtener_valores(registro)
    producto = DynamicProductWrapper(registro, valores)

    # Variable de error para la template
    error = None

    if request.method == 'POST':
        confirmar = request.POST.get('confirmar')

        if confirmar == 'si':
            nombre_producto = producto.nombre

            try:
                # Verificar integridad referencial:
                # Buscar si existen ventas dinámicas que referencien este producto
                ventas_asociadas = _tiene_ventas_asociadas(producto_id)

                if ventas_asociadas:
                    messages.error(
                        request,
                        f'No se puede eliminar "{nombre_producto}" porque tiene ventas registradas.'
                    )
                    return redirect('productos')

                # Eliminar el registro (cascade a ValorCampo)
                DS.eliminar(registro)
                messages.success(
                    request,
                    f'Producto "{nombre_producto}" eliminado correctamente.'
                )
                return redirect('productos')

            except Exception as e:
                logger.exception(f'Error eliminando producto #{producto_id}: {e}')
                messages.error(request, 'No se pudo eliminar el producto. Intenta de nuevo.')
                return redirect('productos')

        else:
            error = 'Debes confirmar la eliminación para continuar.'

    return render(request, 'productos/eliminar_producto_dinamico.html', {
        'producto': producto,
        'error': error,
        'es_admin': es_administrador(request.user),
        'rol_usuario': rol_usuario(request.user),
    })


def _tiene_ventas_asociadas(producto_id):
    """
    Verifica si un producto dinámico tiene ventas asociadas en el formulario Ventas.

    Las ventas dinámicas referencian productos mediante el campo 'producto'
    del formulario 'Ventas', cuyo valor almacena el ID del Registro del producto.

    Args:
        producto_id: ID del Registro del producto a verificar.

    Returns:
        True si existen ventas que referencian este producto.
    """
    try:
        form_ventas = Formulario.objects.get(nombre='Ventas')
    except Formulario.DoesNotExist:
        return False

    campo_producto = Campo.objects.filter(
        formulario=form_ventas,
        nombre='producto',
        activo=True
    ).first()

    if not campo_producto:
        return False

    return ValorCampo.objects.filter(
        campo=campo_producto,
        valor=str(producto_id)
    ).exists()


# ======================================================================
# CATEGORÍAS — Gestión de Opciones Dinámicas
# ======================================================================


@login_required(login_url='login')
@admin_required
def agregar_categoria(request):
    """
    Reemplazo dinámico de la vista legacy agregar_categoria.
    Gestiona las opciones del campo 'categoria' (tipo lista) del formulario Productos
    en lugar del modelo Categoria legacy.
    """
    from django import forms as django_forms

    formulario = DS.obtener_formulario(FORM_PRODUCTOS)
    campo_categoria = formulario.campos.filter(activo=True, nombre='categoria').first()
    opciones = list(campo_categoria.opciones or [])

    class _CategoriaForm(django_forms.Form):
        nombre = django_forms.CharField(
            max_length=100,
            widget=django_forms.TextInput(attrs={
                'class': 'form-control',
                'placeholder': 'Ej: Blusas, Vestidos, Accesorios'
            }),
            label='Nombre de la categoría'
        )

    if request.method == 'POST':
        form = _CategoriaForm(request.POST)
        if form.is_valid():
            nombre = form.cleaned_data['nombre'].strip()
            if nombre not in opciones:
                opciones.append(nombre)
                campo_categoria.opciones = opciones
                campo_categoria.save(update_fields=['opciones'])
                messages.success(request, f'Categoría "{nombre}" agregada correctamente.')
            else:
                messages.warning(request, f'La categoría "{nombre}" ya existe.')
            return redirect('agregar_categoria')
        messages.error(request, 'No se pudo agregar la categoría. Revisa los datos ingresados.')
    else:
        form = _CategoriaForm()

    from types import SimpleNamespace
    categorias = [SimpleNamespace(nombre=op) for op in opciones]

    return render(request, 'formularios/agregar_categoria.html', {
        'form': form,
        'categorias': categorias,
    })


@login_required(login_url='login')
@admin_required
@require_POST
def crear_categoria(request):
    """
    Reemplazo dinámico de la vista legacy crear_categoria (AJAX).
    Agrega una opción al campo 'categoria' del formulario Productos.
    """
    import json
    try:
        data = json.loads(request.body.decode('utf-8') or '{}')
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Solicitud inválida.'
        }, status=400)

    nombre = data.get('nombre', '').strip()
    if not nombre:
        return JsonResponse({
            'success': False,
            'error': 'El nombre es obligatorio.'
        }, status=400)

    formulario = DS.obtener_formulario(FORM_PRODUCTOS)
    campo_categoria = formulario.campos.filter(activo=True, nombre='categoria').first()
    opciones = list(campo_categoria.opciones or [])

    creada = nombre not in opciones
    if creada:
        opciones.append(nombre)
        campo_categoria.opciones = opciones
        campo_categoria.save(update_fields=['opciones'])

    nuevo_id = opciones.index(nombre) + 1

    return JsonResponse({
        'success': True,
        'id': nuevo_id,
        'nombre': nombre,
        'creada': creada,
    })


def _entero(valor, default=0):
    try:
        return int(float(str(valor).replace(',', '.')))
    except (ValueError, TypeError):
        return default


# ======================================================================
# CATÁLOGO PÚBLICO
# ======================================================================


def catalogo_publico(request):
    from apps.shared.configuracion.models import ConfiguracionTienda

    configuracion = ConfiguracionTienda.obtener()
    stock_minimo = configuracion.stock_minimo_alerta
    telefono_whatsapp = ''.join(
        c for c in configuracion.telefono if c.isdigit()
    ) or '573001234567'

    registros = Registro.objects.filter(
        formulario=DS.obtener_formulario(FORM_PRODUCTOS)
    ).order_by('-fecha_creacion')

    valores_map = DS.cargar_valores_mapa(registros)
    productos = [
        DynamicProductWrapper(r, valores_map.get(r.id, {}))
        for r in registros
    ]

    if not configuracion.mostrar_agotados_catalogo:
        productos = [p for p in productos if _entero(p.stock, 0) > 0]

    productos.sort(key=lambda p: (p.nombre or '').lower())

    return render(request, 'public/catalogo.html', {
        'productos': productos,
        'stock_minimo_alerta': stock_minimo,
        'configuracion': configuracion,
        'telefono_whatsapp': telefono_whatsapp,
    })
