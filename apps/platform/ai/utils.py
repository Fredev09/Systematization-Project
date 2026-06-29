"""
utils.py — Utility functions for the AI module.

Helpers for document parsing, content hashing, data extraction,
and JSON-safe serialization for Django session storage.
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import logging
import re
from datetime import datetime, date
from decimal import Decimal
from pathlib import Path, PurePath
from typing import Any, Optional
from uuid import UUID

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def make_json_serializable(obj: Any, _seen: Optional[set] = None) -> Any:
    """
    Recursively convert an object tree to JSON-safe types for session storage.

    Rules:
      - None / str / int / float / bool → returned as-is.
      - list / tuple → recursively serialized list.
      - dict → recursively serialized keys and values.
      - dataclass → fields dict via __dataclass_fields__.
      - Path / PurePath → str().
      - datetime / date → .isoformat().
      - UUID → str().
      - Decimal → float().
      - Objects with .to_dict() → result of .to_dict() (serialized).
      - Django Model instances → pk int (via .pk).
      - bytes → str() (utf-8 decode, fallback to repr).
      - Anything else → str() as last resort.
    """
    # Cycle detection — track objects by id()
    _seen = _seen if _seen is not None else set()
    obj_id = id(obj)
    if obj_id in _seen:
        return f"<circular: {type(obj).__name__}>"
    _seen.add(obj_id)
    try:
        if obj is None or isinstance(obj, (str, int, float, bool)):
            return obj

        if isinstance(obj, (list, tuple)):
            return [make_json_serializable(item, _seen) for item in obj]

        if isinstance(obj, dict):
            return {make_json_serializable(k, _seen): make_json_serializable(v, _seen) for k, v in obj.items()}

        # Dataclass (check before generic hasattr)
        if hasattr(obj, '__dataclass_fields__'):
            return {
                f.name: make_json_serializable(getattr(obj, f.name), _seen)
                for f in obj.__dataclass_fields__.values()
            }

        # Path-like
        if isinstance(obj, (Path, PurePath)):
            return str(obj)

        # datetime / date
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()

        # UUID
        if isinstance(obj, UUID):
            return str(obj)

        # Decimal
        if isinstance(obj, Decimal):
            return float(obj)

        # Objects with to_dict()
        if hasattr(obj, 'to_dict') and callable(obj.to_dict):
            return make_json_serializable(obj.to_dict(), _seen)

        # Django Model instances (has .pk)
        if hasattr(obj, '_meta') and hasattr(obj, 'pk'):
            return obj.pk

        # bytes
        if isinstance(obj, bytes):
            try:
                return obj.decode('utf-8')
            except (UnicodeDecodeError, UnicodeError):
                return repr(obj)

        # Last resort
        return str(obj)
    finally:
        _seen.discard(obj_id)


def extract_text_from_excel(file_path: str | Path, max_rows: int = 100) -> str:
    """
    Extract readable text from an Excel file for AI analysis.
    Returns a text representation: headers + first rows.
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return ""

    lines: list[str] = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx > max_rows:
            break
        row_str = [str(c).strip() if c is not None else "" for c in row]
        row_str = [s for s in row_str if s]
        if row_str:
            lines.append(" | ".join(row_str))

    wb.close()
    return "\n".join(lines)


def extract_text_from_csv(file_path: str | Path, max_rows: int = 100) -> str:
    """
    Extract readable text from a CSV file for AI analysis.
    """
    lines: list[str] = []
    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader):
            if row_idx > max_rows:
                break
            row_str = [s.strip() for s in row if s.strip()]
            if row_str:
                lines.append(" | ".join(row_str))
    return "\n".join(lines)


def extract_text_from_pdf(file_path: str | Path, max_pages: int = 5) -> str:
    """
    Extract text from a PDF. Falls back gracefully if PyMuPDF is not available.
    """
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(file_path)
        text_parts: list[str] = []
        for page_num in range(min(len(doc), max_pages)):
            page = doc[page_num]
            text_parts.append(page.get_text())
        doc.close()
        return "\n\n".join(text_parts)
    except ImportError:
        logger.warning("PyMuPDF not installed. PDF text extraction not available.")
        return f"[PDF file: {Path(file_path).name}] (install PyMuPDF for text extraction)"
    except Exception as e:
        logger.warning(f"PDF extraction error: {e}")
        return f"[PDF file: {Path(file_path).name}] (extraction error: {e})"


def file_to_base64(file_path: str | Path, max_size_mb: int = 20) -> tuple[str, str]:
    """
    Read a file and return (mime_type, base64_data).
    Raises ValueError if file exceeds max_size_mb.
    """
    import base64

    path = Path(file_path)
    suffix = path.suffix.lower()

    file_size = path.stat().st_size
    if file_size > max_size_mb * 1024 * 1024:
        raise ValueError(
            f"File size ({file_size / 1024 / 1024:.1f} MB) exceeds maximum "
            f"of {max_size_mb} MB for base64 encoding."
        )

    mime_map = {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".gif": "image/gif",
        ".webp": "image/webp",
        ".pdf": "application/pdf",
    }
    mime = mime_map.get(suffix, "application/octet-stream")

    with open(path, "rb") as f:
        data = base64.b64encode(f.read()).decode("utf-8")

    return mime, data


def compute_content_hash(content: str | bytes) -> str:
    """Compute SHA256 hash of content for caching."""
    if isinstance(content, str):
        content = content.encode("utf-8")
    return hashlib.sha256(content).hexdigest()


def compute_file_hash(file_path: str | Path) -> str:
    """Compute SHA256 hash of a file for caching."""
    sha = hashlib.sha256()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            sha.update(chunk)
    return sha.hexdigest()


def safe_json_parse(text: str) -> Optional[dict[str, Any]]:
    """
    Safely parse JSON from AI text response.
    Handles markdown code blocks and trailing text.
    """
    if not text:
        return None

    text = text.strip()

    # Remove markdown code block delimiters
    if text.startswith("```"):
        for delim in ["```json\n", "```json", "```\n", "```"]:
            if delim in text:
                text = text.split(delim, 1)[-1]
                if text.endswith("```"):
                    text = text[:-3]
                break

    text = text.strip()

    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    # Try to find JSON object in the text
    json_match = re.search(r"\{.*\}", text, re.DOTALL)
    if json_match:
        try:
            return json.loads(json_match.group())
        except json.JSONDecodeError:
            pass

    return None


def truncate_text(text: str, max_chars: int = 10000) -> str:
    """Truncate text to max_chars, keeping complete words."""
    if len(text) <= max_chars:
        return text
    return text[:max_chars].rsplit(" ", 1)[0] + "\n... [truncated]"


def guess_field_type_from_value(value: str) -> str:
    """
    Try to guess the field type from a sample value.
    Returns a FieldType value or 'texto' as fallback.
    """
    if not value or not value.strip():
        return "texto"

    v = value.strip()

    # Email
    if re.match(r"^[\w.+-]+@[\w-]+\.[\w.]+$", v):
        return "email"

    # URL
    if v.startswith(("http://", "https://", "www.")):
        return "url"

    # Phone (Colombian format)
    if re.match(r"^\+?3?\d{7,10}$", v) or re.match(r"^\d{3}[- ]?\d{3}[- ]?\d{4}$", v):
        return "telefono"

    # Date
    if re.match(r"^\d{4}-\d{2}-\d{2}$", v) or re.match(r"^\d{2}/\d{2}/\d{4}$", v):
        return "fecha"

    # Boolean
    if v.lower() in ("si", "no", "sí", "true", "false", "1", "0", "x", "✓", "✔"):
        return "booleano"

    # Number / Currency
    clean = v.replace("$", "").replace(",", "").replace(".", "").strip()
    if clean.lstrip("-").isdigit():
        if "." in v or "," in v or "$" in v:
            return "moneda"
        return "numero"

    # Decimal number
    try:
        float(v.replace(",", "."))
        return "numero"
    except ValueError:
        pass

    return "texto"
