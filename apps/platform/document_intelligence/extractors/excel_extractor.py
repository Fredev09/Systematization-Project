"""
excel_extractor.py — Excel document extractor.

Extracts all sheets, tables, columns, rows, and metadata from .xlsx files.
Reuses openpyxl (already in project dependencies).
"""

from __future__ import annotations

import logging
from pathlib import Path
from datetime import datetime as dt_lib
from typing import Any, Optional

from openpyxl import load_workbook

from apps.platform.document_intelligence.extractors.base import (
    DocumentExtractor,
    ExtractedDocument,
    ExtractedTable,
)

logger = logging.getLogger(__name__)

# Max rows to extract per sheet
MAX_ROWS_PER_SHEET = 50000


class ExcelExtractor(DocumentExtractor):
    """Extracts content from Excel (.xlsx) files."""

    def extract(self, file_path: str | Path) -> ExtractedDocument:
        path = Path(file_path)
        doc = ExtractedDocument(
            document_type="excel",
            title=path.name,
        )

        # .xls is NOT supported by openpyxl
        if path.suffix.lower() == ".xls":
            raise ValueError(
                "The file format is .xls (legacy Excel). "
                "Only .xlsx files are supported. Please save the file as .xlsx and try again."
            )

        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            if "encrypted" in str(e).lower():
                raise ValueError("The Excel file is password-protected. Remove protection first.")
            raise ValueError(f"Cannot read Excel file: {e}")

        doc.metadata["total_sheets"] = len(wb.sheetnames)
        doc.sheets = wb.sheetnames

        all_text_parts: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_tables = self._extract_sheet(ws, sheet_name)

            for table in sheet_tables:
                doc.tables.append(table)
                # Build text representation
                all_text_parts.append(f"=== Sheet: {sheet_name} ===")
                if table.headers:
                    all_text_parts.append(" | ".join(table.headers))
                for row in table.rows[:100]:  # First 100 rows for text
                    all_text_parts.append(" | ".join(row))

            # If it's the first sheet with data, use as primary
            if sheet_tables and not doc.columns:
                primary = sheet_tables[0]
                doc.columns = primary.headers
                doc.rows = primary.rows

        wb.close()
        doc.raw_text = "\n".join(all_text_parts)
        doc.confidence = 1.0 if doc.columns else 0.5
        doc.metadata["total_rows"] = doc.total_rows
        doc.metadata["total_columns"] = doc.total_columns

        return doc

    def _extract_sheet(self, ws, sheet_name: str) -> list[ExtractedTable]:
        """Extract tables from a single worksheet (iterator-based, OOM-safe)."""
        row_iter = ws.iter_rows(values_only=True)

        headers: list[str] = []
        header_row_idx = 0

        # Scan up to 20 rows to detect headers
        preview_rows: list[tuple] = []
        for idx, row in enumerate(row_iter):
            if row is None:
                continue
            preview_rows.append(row)
            if not headers and any(c is not None for c in row):
                candidate = [str(c).strip() if c is not None else "" for c in row]
                valid_headers = [h for h in candidate if h]
                if len(valid_headers) >= 2:
                    headers = candidate
                    header_row_idx = idx
                    break
            if idx >= 19:
                break

        if not headers:
            return []

        # Continue reading from where the iterator left off
        cleaned_rows: list[list[str]] = []
        row_count = 0
        for row in row_iter:
            if row is None:
                continue
            row_count += 1
            if row_count > MAX_ROWS_PER_SHEET:
                logger.warning("Sheet '%s' truncated at %d rows", sheet_name, MAX_ROWS_PER_SHEET)
                break
            row_str = [self._cell_to_str(c) for c in row]
            if any(c for c in row_str):
                cleaned_rows.append(row_str)

        table = ExtractedTable(
            name=sheet_name,
            headers=headers,
            rows=cleaned_rows,
            row_count=row_count,
            confidence=1.0,
        )

        return [table]

    @staticmethod
    def _cell_to_str(cell: Any) -> str:
        """Convert a cell value to string."""
        if cell is None:
            return ""
        if isinstance(cell, dt_lib):
            return cell.strftime("%Y-%m-%d")
        if isinstance(cell, (int, float)):
            if isinstance(cell, float) and cell == int(cell):
                return str(int(cell))
            # Format: remove trailing zeros after decimal
            s = f"{cell:.10f}".rstrip("0").rstrip(".")
            return s
        return str(cell).strip()

    @classmethod
    def supported_extensions(cls) -> list[str]:
        return [".xlsx", ".xls"]
