"""
ExcelImportService — Importación de Excel a Dynamic Forms.

Reutiliza al máximo las validaciones y creación del sistema existente:
- DynamicService.validar_completo() para pre-validación sin escribir BD.
- DynamicService.crear() para escritura.
- DynamicService.actualizar() para actualización.
- DynamicService.upsert_por_identificador() para upsert.

Módulo de matching delegado a column_matching.ColumnMatcher.
No duplica lógica de normalización ni matching.
"""

from __future__ import annotations

import logging
import time
from collections import Counter, defaultdict
from datetime import datetime as dt_lib
from io import BytesIO
from pathlib import Path
from typing import Any, Optional

from django.db import transaction
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .column_matching import (
    ColumnMatchResult,
    ColumnMatcher,
    normalizar_columna,
)
from .services_dynamic import DynamicService as DS
from .validators import (
    normalizar_booleano,
    normalizar_fecha,
    normalizar_moneda,
    normalizar_numero,
)

logger = logging.getLogger(__name__)

# ======================================================================
# DIAGNÓSTICO — Helpers de instrumentación temporal
# ======================================================================
def _diag_log(stage: str, start: float, detail: str = ''):
    """Registra duración de un paso. Warning si >500ms."""
    elapsed_ms = (time.perf_counter() - start) * 1000
    msg = f'[DIAG] {stage}: {elapsed_ms:.1f}ms'
    if detail:
        msg += f' | {detail}'
    if elapsed_ms > 500:
        logger.warning(f'[DIAG] ⚠ {stage} EXCEDE 500ms: {elapsed_ms:.1f}ms | {detail}')
    else:
        logger.info(msg)
    return time.perf_counter()


def _diag_start(stage: str):
    """Marca inicio de una etapa de diagnóstico."""
    logger.info(f'[DIAG] >>> {stage}')
    return time.perf_counter()

# ======================================================================
# CONSTANTES
# ======================================================================

MAX_IMPORT_FILE_SIZE: int = 50 * 1024 * 1024  # 50 MB
"""Tamaño máximo permitido para archivos de importación."""

MAX_IMPORT_ROWS: int = 50000
"""Máximo de filas de datos permitidas en una importación."""

EXTENSIONES_EXCEL_PERMITIDAS: set[str] = {'.xlsx'}
"""Extensiones de archivo permitidas para importación."""

MIMETYPES_EXCEL: set[str] = {
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    'application/octet-stream',  # fallback común en algunos navegadores
}
"""MIME types aceptados para archivos Excel."""

MODOS_IMPORTACION: dict[str, str] = {
    'crear': 'Solo crear registros nuevos',
    'actualizar': 'Solo actualizar registros existentes',
    'upsert': 'Crear o actualizar (UPSERT)',
    'validar': 'Solo validar (Dry Run)',
}

TIPOS_EXCLUIDOS_MAPEO: set[str] = {'calculado', 'imagen', 'archivo'}

# Cache de estilos Excel para no recrearlos en cada llamada
_EXCEL_STYLES_CACHE: dict[str, Any] = {}


def _sanitizar_extension(nombre_archivo: str) -> str:
    """
    Valida y retorna la extensión de un archivo si está en la lista blanca.
    Previene path traversal y extensiones maliciosas.
    """
    ext = Path(nombre_archivo).suffix.lower()
    if ext not in EXTENSIONES_EXCEL_PERMITIDAS:
        raise ValueError(
            f'Extensión "{ext}" no permitida. Solo se aceptan: '
            f'{", ".join(sorted(EXTENSIONES_EXCEL_PERMITIDAS))}'
        )
    return ext


def _validar_archivo_importacion(archivo) -> None:
    """
    Valida seguridad del archivo subido: tamaño, extensión y contenido.
    Lanza ValueError si alguna validación falla.
    """
    if archivo.size > MAX_IMPORT_FILE_SIZE:
        tam_mb = archivo.size / (1024 * 1024)
        raise ValueError(
            f'El archivo excede el tamaño máximo de '
            f'{MAX_IMPORT_FILE_SIZE // (1024 * 1024)} MB '
            f'({tam_mb:.1f} MB recibidos).'
        )

    if not archivo.name:
        raise ValueError('El archivo no tiene nombre.')

    _sanitizar_extension(archivo.name)

    content_type = getattr(archivo, 'content_type', '')
    if content_type and content_type not in MIMETYPES_EXCEL:
        logger.warning(
            f'MIME type inesperado para "{archivo.name}": {content_type}'
        )


def _limitar_filas(filas: list[dict], max_filas: int = MAX_IMPORT_ROWS) -> list[dict]:
    """Limita el número de filas a procesar y emite advertencia si se excede."""
    if len(filas) > max_filas:
        logger.warning(
            f'Archivo excede el límite de {max_filas} filas '
            f'({len(filas)} encontradas). Procesando solo las primeras {max_filas}.'
        )
        return filas[:max_filas]
    return filas


def _estilos_excel() -> dict[str, Any]:
    """Retorna diccionario de estilos Excel cacheados (FASE 2: rendimiento)."""
    if not _EXCEL_STYLES_CACHE:
        _EXCEL_STYLES_CACHE.update({
            'titulo_fill': PatternFill('solid', fgColor='D41473'),
            'encabezado_fill': PatternFill('solid', fgColor='FCE7F3'),
            'ejemplo_fill': PatternFill('solid', fgColor='FFF1F8'),
            'error_fill': PatternFill('solid', fgColor='FEF2F2'),
            'titulo_font': Font(color='FFFFFF', bold=True, size=14),
            'encabezado_font': Font(color='111827', bold=True, size=11),
            'ejemplo_font': Font(color='4B5563', italic=True, size=10),
            'comentario_font': Font(color='6B7280', italic=True, size=9),
            'error_font': Font(color='DC2626'),
            'moneda_format': '#,##0',
            'fecha_format': 'YYYY-MM-DD',
        })
    return _EXCEL_STYLES_CACHE


# ======================================================================
# Normalización de valores (FASE 5: Importación robusta)
# ======================================================================

_NORMALIZADORES: dict[str, callable] = {
    'booleano': normalizar_booleano,
    'fecha': normalizar_fecha,
    'moneda': normalizar_moneda,
    'numero': lambda v: normalizar_numero(v, permitir_decimales=True),
}


def _normalizar_valor_importacion(tipo_campo: str, valor_raw: str) -> str:
    """
    Normaliza un valor según el tipo de campo antes de la validación.
    Usa los normalizadores de validators.py.

    Esto permite que formatos como "$50,000", "1.000,50", "X", "✓",
    "DD/MM/YYYY" sean convertidos automáticamente.
    """
    if not valor_raw:
        return valor_raw
    normalizador = _NORMALIZADORES.get(tipo_campo)
    if normalizador:
        resultado = normalizador(valor_raw)
        if resultado is not None:
            return resultado
    return valor_raw


# ======================================================================
# Conversión de valor de celda (FASE 5: fechas seriales, monedas)
# ======================================================================


def _valor_celda(cell, cell_coord: str = '') -> str:
    """Convierte una celda de openpyxl a string plano con normalización.

    Args:
        cell: Valor de la celda (None, datetime, int, float, str, etc.).
        cell_coord: Coordenada de la celda (ej: 'A1') para diagnóstico.

    Returns:
        String normalizado. Vacío si la celda está vacía o no tiene valor.
    """
    if cell is None:
        return ''

    if isinstance(cell, dt_lib):
        return cell.strftime('%Y-%m-%d')

    if isinstance(cell, (int, float)):
        if isinstance(cell, float) and cell == int(cell):
            return str(int(cell))
        return str(cell)

    valor = str(cell).strip()
    # Detectar si es una fórmula sin caché (comienza con =)
    if valor.startswith('='):
        logger.warning(
            f'Celda {cell_coord} contiene fórmula sin valor calculado: "{valor[:100]}". '
            'El archivo debe guardarse con valores calculados.'
        )
        return ''

    return valor


def _excluir_tipo(campo) -> bool:
    """Determina si un campo debe excluirse del mapeo."""
    return campo.tipo in TIPOS_EXCLUIDOS_MAPEO if hasattr(campo, 'tipo') else False


# ======================================================================
# Parseo de filas de datos
# ======================================================================


def parse_data_rows(
    ws, header_row_idx: int, headers: list[str],
    data_start_row: Optional[int] = None
) -> list[dict]:
    """
    Parsea las filas de datos a partir de una hoja y fila de encabezados.

    Maneja:
      - Filas completamente vacías (se skipean)
      - Celdas combinadas (merged cells)
      - Fórmulas sin valor cachead
      - Celdas con fecha serial
    """
    _t0 = time.perf_counter()
    if data_start_row is None:
        data_start_row = header_row_idx + 1

    indices_validos = [i for i, h in enumerate(headers) if h]
    headers_filtrados = [headers[i] for i in indices_validos]

    _t_iter = _diag_start('parse_data_rows: iter_rows')
    filas_data = list(ws.iter_rows(values_only=True))
    _diag_log('parse_data_rows: iter_rows', _t_iter, f'total_en_hoja={len(filas_data)}, data_start_row={data_start_row}')
    data_rows = filas_data[data_start_row:] if data_start_row < len(filas_data) else []

    from openpyxl.utils import get_column_letter

    _t_loop = _diag_start(f'parse_data_rows: loop over {len(data_rows)} filas')
    filas = []
    for row_idx, row in enumerate(data_rows):
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue
        fila_dict = {}
        for i in indices_validos:
            val = row[i] if i < len(row) else None
            col_letter = get_column_letter(i + 1)
            actual_row = data_start_row + row_idx + 1
            cell_coord = f'{col_letter}{actual_row}'
            fila_dict[headers_filtrados[indices_validos.index(i)]] = _valor_celda(val, cell_coord)
        filas.append(fila_dict)

    _diag_log('parse_data_rows: loop', _t_loop, f'filas_output={len(filas)}')
    logger.info(f'[DIAG] <<< parse_data_rows total: {(time.perf_counter() - _t0) * 1000:.1f}ms')
    return filas


# ======================================================================
# Lectura de Excel (backward compatible)
# ======================================================================


_EXCEL_ENCRYPTED_MSG = 'The workbook is encrypted'
"""Mensaje de error de openpyxl para archivos protegidos con contraseña."""


def leer_excel(archivo):
    """
    Lee un archivo .xlsx y retorna sus datos como estructura plana.
    Mantiene compatibilidad hacia atrás.

    Maneja:
      - Archivos protegidos con contraseña
      - Celdas con fórmulas sin caché
      - Celdas combinadas (merged cells)
      - Archivos corruptos
    """
    try:
        wb = load_workbook(archivo, read_only=True, data_only=True)
    except Exception as e:
        err_str = str(e)
        if _EXCEL_ENCRYPTED_MSG in err_str:
            raise ValueError(
                'El archivo Excel está protegido con contraseña. '
                'Debes eliminar la protección antes de importar.'
            )
        raise ValueError(f'No se pudo leer el archivo Excel: {e}')

    try:
        ws = wb.active
        if ws is None:
            raise ValueError('El archivo Excel no tiene hojas de cálculo.')

        filas_iter = ws.iter_rows(values_only=True)

        try:
            raw_headers = next(filas_iter)
        except StopIteration:
            raise ValueError('El archivo Excel está vacío.')

        if not raw_headers:
            raise ValueError('La primera fila (encabezados) está vacía.')

        encabezados = [str(h).strip() if h is not None else '' for h in raw_headers]

        if not any(encabezados):
            raise ValueError('No se encontraron encabezados de columna en la primera fila.')

        filas = parse_data_rows(ws, 0, encabezados)

        if not filas:
            raise ValueError('El archivo Excel no contiene datos (solo encabezados).')

    finally:
        wb.close()

    return encabezados, filas


# ======================================================================
# Detección y corrección de mapeo de columnas
# ======================================================================


def detectar_columnas(encabezados, formulario):
    """Detecta automáticamente la correspondencia entre columnas y campos."""
    matcher = ColumnMatcher(formulario)
    results = matcher.match_all(encabezados)
    return {r.column_index: r.matched_to for r in results if r.matched_to}


def construir_mapeo_completo(encabezados, formulario, mapeo_usuario=None):
    """Construye el mapeo combinando detección automática y correcciones."""
    matcher = ColumnMatcher(formulario)
    mapeo, sin_mapear, _results = matcher.build_mapping(
        encabezados, user_overrides=mapeo_usuario
    )
    sin_mapear_filtrados = [
        nombre for nombre in sin_mapear
        if not _excluir_tipo(formulario.campos.filter(nombre=nombre).first())
    ]
    return mapeo, sin_mapear_filtrados


# ======================================================================
# Validación avanzada (pre-preview)
# ======================================================================


def _check_columnas_duplicadas(encabezados: list[str], resultado: dict) -> None:
    """Detecta columnas con el mismo nombre en el Excel."""
    contador_columnas = Counter(h.lower() for h in encabezados if h)
    duplicadas = [nombre for nombre, count in contador_columnas.items() if count > 1]
    if duplicadas:
        resultado['columnas_duplicadas'] = duplicadas
        resultado['advertencias'].append(
            f'Columnas duplicadas en el Excel: {", ".join(duplicadas)}. '
            'Se usará la primera coincidencia.'
        )


def _check_columnas_vacias(encabezados: list[str], filas: list[dict], resultado: dict) -> None:
    """Detecta columnas sin datos en ninguna fila."""
    if not filas:
        return
    for idx, header in enumerate(encabezados):
        if not header:
            continue
        if all(not fila.get(header, '') for fila in filas):
            resultado['columnas_vacias'].append(idx)


def _check_columnas_desconocidas(match_results: list | None, resultado: dict) -> None:
    """Detecta columnas del Excel sin correspondencia en el formulario."""
    for r in match_results or []:
        if r.column_name and not r.matched_to:
            resultado['columnas_desconocidas'].append((r.column_index, r.column_name))


def _check_campos_obligatorios_faltantes(
    formulario, mapeo: dict, resultado: dict
) -> None:
    """Detecta campos obligatorios del formulario no mapeados."""
    campos_mapeados = set(mapeo.values())
    for campo in formulario.campos.filter(activo=True, obligatorio=True):
        if not _excluir_tipo(campo) and campo.nombre not in campos_mapeados:
            resultado['errores'].append(
                f'El campo obligatorio "{campo.nombre}" no está mapeado a ninguna columna.'
            )


def _check_identificadores_repetidos(
    formulario, encabezados: list[str], filas: list[dict], mapeo: dict, resultado: dict
) -> None:
    """Detecta valores duplicados del identificador principal en el Excel."""
    campo_id = DS.obtener_identificador_principal(formulario.nombre)
    if not campo_id or campo_id.nombre not in set(mapeo.values()):
        return

    col_idx_id = next(
        (idx for idx, name in mapeo.items() if name == campo_id.nombre),
        None
    )
    if col_idx_id is None or col_idx_id >= len(encabezados):
        return

    header_id = encabezados[col_idx_id]
    valores_id = [
        fila.get(header_id, '').strip()
        for fila in filas if fila.get(header_id, '').strip()
    ]
    contador_ids = Counter(valores_id)
    repetidos = [val for val, count in contador_ids.items() if count > 1]
    if repetidos:
        resultado['ids_repetidos'] = repetidos[:10]
        resultado['advertencias'].append(
            f'Identificadores repetidos en el Excel: {", ".join(repetidos[:5])}'
            + (f' y {len(repetidos) - 5} más.' if len(repetidos) > 5 else '.')
        )


def _check_filas_duplicadas(filas: list[dict], resultado: dict) -> None:
    """Detecta filas completamente duplicadas en el Excel."""
    if not filas:
        return
    tuplas_filas = [tuple(sorted(fila.items())) for fila in filas]
    contador_filas = Counter(tuplas_filas)
    duplicados = sum(count - 1 for count in contador_filas.values() if count > 1)
    resultado['filas_duplicadas'] = duplicados
    if duplicados > 0:
        resultado['advertencias'].append(
            f'{duplicados} fila(s) duplicada(s) en el Excel. '
            'Solo se importará la primera ocurrencia de cada una.'
        )


def validar_estructura(formulario, encabezados, filas, mapeo, match_results) -> dict:
    """
    Validación avanzada de estructura del Excel antes del preview.

    Detecta:
      - Identificadores repetidos dentro del Excel
      - Filas completamente duplicadas
      - Columnas duplicadas, vacías, desconocidas
      - Campos obligatorios faltantes
    """
    resultado: dict = {
        'valido': True,
        'advertencias': [],
        'errores': [],
        'columnas_duplicadas': [],
        'columnas_vacias': [],
        'columnas_desconocidas': [],
        'ids_repetidos': [],
        'filas_duplicadas': 0,
    }

    _check_columnas_duplicadas(encabezados, resultado)
    _check_columnas_vacias(encabezados, filas, resultado)
    _check_columnas_desconocidas(match_results, resultado)
    _check_campos_obligatorios_faltantes(formulario, mapeo, resultado)
    _check_identificadores_repetidos(formulario, encabezados, filas, mapeo, resultado)
    _check_filas_duplicadas(filas, resultado)

    resultado['valido'] = len(resultado['errores']) == 0
    return resultado


# ======================================================================
# Pre-visualización (validación sin escribir BD)
# ======================================================================


def previsualizar(formulario, encabezados, filas, mapeo):
    """
    Valida todas las filas del Excel contra el formulario destino
    SIN escribir en la base de datos.

    Returns:
        list[dict] — una entrada por fila:
            - fila_idx: int (0-based)
            - valores: dict {nombre_campo: valor}
            - valida: bool
            - errores: list[str]
    """
    resultados: list[dict] = []

    for fila_idx, fila in enumerate(filas):
        valores_dict: dict[str, str] = {}

        for col_idx, nombre_campo in mapeo.items():
            if col_idx < len(encabezados):
                enc = encabezados[col_idx]
                valores_dict[nombre_campo] = fila.get(enc, '')

        errores = DS.validar_completo(formulario, valores_dict)

        resultados.append({
            'fila_idx': fila_idx,
            'valores': valores_dict,
            'valida': len(errores) == 0,
            'errores': errores,
        })

    return resultados


# ======================================================================
# Procesamiento de fila por modo (extraído para claridad, FASE 9)
# ======================================================================


def _process_row_for_mode(
    modo: str, row: dict, formulario, usuario,
    campo_id_nombre: Optional[str],
    campo_id,
) -> dict:
    """
    Procesa una fila individual según el modo de importación.

    Returns:
        {'accion': 'creado'|'actualizado'|'ignorado', ...}
    """
    if modo == 'validar':
        return {'accion': 'ignorado'}

    elif modo == 'crear':
        DS.crear(formulario.nombre, row['valores'], usuario=usuario)
        return {'accion': 'creado'}

    elif modo == 'actualizar':
        if not campo_id_nombre:
            raise ValueError(
                'El formulario no tiene un identificador principal configurado. '
                'No se puede usar el modo "Actualizar".'
            )
        valor_id = row['valores'].get(campo_id_nombre, '').strip()
        if not valor_id:
            return {'accion': 'ignorado'}

        registro_existente = DS.buscar_por_identificador(
            formulario.nombre, valor_id
        )
        if registro_existente:
            DS.actualizar(registro_existente, row['valores'], usuario=usuario)
            return {'accion': 'actualizado'}
        else:
            return {'accion': 'ignorado'}

    elif modo == 'upsert':
        if not campo_id_nombre:
            raise ValueError(
                'El formulario no tiene un identificador principal configurado. '
                'No se puede usar el modo "UPSERT".'
            )
        _registro, fue_creado = DS.upsert_por_identificador(
            formulario.nombre,
            row['valores'],
            usuario=usuario,
        )
        return {'accion': 'creado' if fue_creado else 'actualizado'}

    return {'accion': 'ignorado'}


# ======================================================================
# Importación definitiva con modos
# ======================================================================


def importar(
    formulario, preview_rows, usuario=None,
    modo: str = 'crear', mapeo=None
) -> dict:
    """
    Importa filas pre-validadas usando DynamicService.

    Args:
        formulario: Formulario destino.
        preview_rows: list[dict] — resultado de previsualizar().
        usuario: User opcional.
        modo: 'crear', 'actualizar', 'upsert', 'validar'.
        mapeo: dict {indice: campo_nombre}.

    Returns:
        dict con total, creados, actualizados, ignorados, errores, tiempo_seg, modo.
    """
    creados = 0
    actualizados = 0
    ignorados = 0
    errores: list[dict] = []
    total = len(preview_rows)
    inicio = time.time()

    campo_id = DS.obtener_identificador_principal(formulario.nombre)
    campo_id_nombre = campo_id.nombre if campo_id else None

    for row in preview_rows:
        try:
            with transaction.atomic():
                resultado = _process_row_for_mode(
                    modo, row, formulario, usuario, campo_id_nombre, campo_id
                )
                accion = resultado['accion']
                if accion == 'creado':
                    creados += 1
                elif accion == 'actualizado':
                    actualizados += 1
                else:
                    ignorados += 1

        except Exception as e:
            logger.exception(f'Error importando fila #{row["fila_idx"] + 1}: {e}')
            errores.append({
                'fila_idx': row['fila_idx'],
                'valores': row['valores'],
                'error': str(e),
            })

    tiempo_total = time.time() - inicio

    return {
        'total': total,
        'creados': creados,
        'actualizados': actualizados,
        'ignorados': ignorados,
        'errores': errores,
        'tiempo_seg': round(tiempo_total, 2),
        'modo': modo,
    }


# ======================================================================
# Generación de Excel de errores
# ======================================================================


def generar_excel_errores(formulario_nombre: str, errores: list[dict]) -> Optional[BytesIO]:
    """
    Genera un archivo Excel descargable con los errores de importación.
    """
    if not errores:
        return None

    estilos = _estilos_excel()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Errores'

    ws.append([f'Errores de importación: {formulario_nombre}'])
    ws.merge_cells('A1:E1')
    ws['A1'].fill = estilos['titulo_fill']
    ws['A1'].font = Font(color='FFFFFF', bold=True, size=14)

    headers = ['Fila', 'Campo', 'Valor recibido', 'Mensaje', 'Sugerencia']
    ws.append([])
    ws.append(headers)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.fill = estilos['encabezado_fill']
        cell.font = estilos['encabezado_font']

    for err in errores:
        valores = err.get('valores', {})
        fila_idx = err.get('fila_idx', 0) + 1
        mensaje = err.get('error', '')
        if valores:
            for campo_nombre, valor in valores.items():
                ws.append([fila_idx, campo_nombre, valor, mensaje, ''])
        else:
            ws.append([fila_idx, '', '', mensaje, ''])

    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.fill = estilos['error_fill']
            cell.font = estilos['error_font']

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 40

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ======================================================================
# Generación de plantilla Excel
# ======================================================================


_ETIQUETAS_TIPO: dict[str, str] = {
    'texto': 'Texto', 'textarea': 'Texto largo', 'numero': 'Número',
    'moneda': 'Ej: 50000', 'fecha': 'Ej: 2024-01-15',
    'booleano': 'Sí / No', 'lista': 'Elegir de lista',
    'email': 'Ej: correo@ejemplo.com', 'telefono': 'Ej: 3001234567',
}

_EJEMPLOS_TIPO: dict[str, str] = {
    'texto': 'Ejemplo', 'textarea': 'Descripción del producto',
    'numero': '10', 'moneda': '50000', 'fecha': '2024-01-15',
    'booleano': 'Sí', 'email': 'correo@ejemplo.com',
    'telefono': '3001234567',
}


def _cargar_campos_plantilla(
    formulario,
) -> tuple[list[str], dict[str, str], dict[str, list[str]], list]:
    """Carga y filtra los campos del formulario para la plantilla."""
    campos = list(formulario.campos.filter(activo=True).order_by('orden'))
    encabezados: list[str] = []
    tipos_campo: dict[str, str] = {}
    opciones_campo: dict[str, list[str]] = {}

    for campo in campos:
        if _excluir_tipo(campo):
            continue
        encabezados.append(campo.nombre)
        tipos_campo[campo.nombre] = campo.tipo
        if campo.tipo == 'lista' and campo.opciones:
            opciones_campo[campo.nombre] = [
                o.strip() for o in campo.opciones.split('\n') if o.strip()
            ]

    return encabezados, tipos_campo, opciones_campo, campos


def _escribir_encabezados_plantilla(ws, estilos, encabezados, formulario):
    """Escribe título, nota y fila de encabezados en la plantilla."""
    num_cols = len(encabezados)
    if num_cols == 0:
        raise ValueError('El formulario no tiene campos mapeables para la plantilla.')

    ws.append([f'Plantilla de importación: {formulario.nombre}'])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    ws['A1'].fill = estilos['titulo_fill']
    ws['A1'].font = Font(color='FFFFFF', bold=True, size=15)

    ws.append(['Los campos marcados con * son obligatorios.'])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)

    ws.append([])
    ws.append(encabezados)

    for col_idx, nombre in enumerate(encabezados, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.fill = estilos['encabezado_fill']
        cell.font = estilos['encabezado_font']

    return num_cols


def _escribir_fila_ayuda(ws, estilos, encabezados, tipos_campo, campos, num_cols):
    """Escribe la fila de ayuda con tipos y obligatoriedad."""
    fila_ayuda = []
    for nombre in encabezados:
        tipo = tipos_campo.get(nombre, 'texto')
        obligatorio = any(c.nombre == nombre and c.obligatorio for c in campos)
        etiqueta = _ETIQUETAS_TIPO.get(tipo, tipo)
        if obligatorio:
            etiqueta += ' *'
        fila_ayuda.append(etiqueta)

    ws.append(fila_ayuda)
    for col_idx in range(1, num_cols + 1):
        ws.cell(row=5, column=col_idx).font = estilos['comentario_font']


def _escribir_fila_ejemplo(ws, estilos, encabezados, tipos_campo, opciones_campo, num_cols):
    """Escribe la fila de ejemplo con valores ilustrativos."""
    fila_ejemplo = []
    for nombre in encabezados:
        tipo = tipos_campo.get(nombre, 'texto')
        if tipo == 'lista':
            opciones = opciones_campo.get(nombre, [])
            ejemplo = opciones[0] if opciones else 'Opción'
        else:
            ejemplo = _EJEMPLOS_TIPO.get(tipo, '')
        fila_ejemplo.append(ejemplo)

    ws.append(fila_ejemplo)
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=6, column=col_idx)
        cell.fill = estilos['ejemplo_fill']
        cell.font = estilos['ejemplo_font']


def _agregar_validaciones_plantilla(ws, encabezados, tipos_campo, opciones_campo, estilos):
    """Agrega validaciones de datos (listas desplegables, booleanos, formatos)."""
    num_cols = len(encabezados)

    for nombre, opciones in opciones_campo.items():
        if nombre in encabezados:
            col_idx_enc = encabezados.index(nombre) + 1
            col_letter = get_column_letter(col_idx_enc)
            dv = DataValidation(
                type='list',
                formula1=f'"{",".join(opciones)}"',
                allow_blank=True,
            )
            dv.error = f'Selecciona un valor de la lista: {", ".join(opciones)}'
            dv.errorTitle = 'Valor inválido'
            dv.prompt = f'Valores válidos: {", ".join(opciones)}'
            dv.promptTitle = nombre
            ws.add_data_validation(dv)
            dv.add(f'{col_letter}7:{col_letter}1048576')

    for nombre in encabezados:
        col_idx_enc = encabezados.index(nombre) + 1
        col_letter = get_column_letter(col_idx_enc)
        tipo = tipos_campo.get(nombre, 'texto')

        if tipo == 'booleano':
            dv = DataValidation(
                type='list',
                formula1='"Sí,No"',
                allow_blank=True,
            )
            dv.error = 'Usa "Sí" o "No"'
            dv.errorTitle = 'Valor booleano inválido'
            ws.add_data_validation(dv)
            dv.add(f'{col_letter}7:{col_letter}1048576')

        elif tipo in ('moneda', 'fecha'):
            formato = estilos['moneda_format'] if tipo == 'moneda' else estilos['fecha_format']
            for row_num in range(7, 1007):
                ws.cell(row=row_num, column=col_idx_enc).number_format = formato

    ws.column_dimensions['A'].width = 18
    for i in range(2, num_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22


def generar_plantilla_excel(formulario) -> BytesIO:
    """
    Genera un archivo Excel plantilla con la definición del formulario.
    Incluye encabezados con comentarios, fila de ejemplo,
    listas desplegables, validación de datos y formato.
    """
    estilos = _estilos_excel()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Plantilla'

    encabezados, tipos_campo, opciones_campo, campos = _cargar_campos_plantilla(formulario)
    num_cols = _escribir_encabezados_plantilla(ws, estilos, encabezados, formulario)
    _escribir_fila_ayuda(ws, estilos, encabezados, tipos_campo, campos, num_cols)
    _escribir_fila_ejemplo(ws, estilos, encabezados, tipos_campo, opciones_campo, num_cols)
    _agregar_validaciones_plantilla(ws, encabezados, tipos_campo, opciones_campo, estilos)

    _crear_hoja_instrucciones(wb, formulario, campos)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _crear_hoja_instrucciones(wb, formulario, campos):
    """Crea hoja de instrucciones en el workbook de la plantilla."""
    ws_instrucciones = wb.create_sheet('Instrucciones')
    ws_instrucciones.append(['INSTRUCCIONES DE IMPORTACIÓN'])
    ws_instrucciones['A1'].font = Font(bold=True, size=14, color='D41473')
    ws_instrucciones.append([])

    instrucciones = [
        '1. Completa los datos en la hoja "Plantilla" a partir de la fila 7.',
        '2. No modifiques los encabezados (fila 4).',
        '3. Los campos con * son obligatorios.',
        '4. Para campos de lista, usa los valores disponibles en el menú desplegable.',
        '5. Para booleanos, usa "Sí" o "No".',
        '6. Para fechas, usa el formato YYYY-MM-DD (ej: 2024-01-15).',
        '7. Para moneda, ingresa solo números sin símbolos (ej: 50000).',
        '',
        'Campos del formulario:',
    ]
    for inst in instrucciones:
        ws_instrucciones.append([inst])

    for campo in campos:
        if not _excluir_tipo(campo):
            info = f'  • {campo.nombre} ({campo.get_tipo_display()})'
            if campo.obligatorio:
                info += ' — OBLIGATORIO'
            if campo.unico:
                info += ' — ÚNICO'
            ws_instrucciones.append([info])

    ws_instrucciones.column_dimensions['A'].width = 80


# ======================================================================
# Análisis completo de workbook
# ======================================================================


def _detectar_mejor_hoja(
    wb, matcher: ColumnMatcher
) -> tuple[Any, str, dict[str, float], int]:
    """Detecta la mejor hoja del workbook por puntaje de coincidencia."""
    _t0 = time.perf_counter()
    sheet_names = wb.sheetnames
    total_sheets = len(sheet_names)
    best_sheet_name: str | None = None
    best_ws = None
    best_score = -1.0
    all_sheets_scores: dict[str, float] = {}

    logger.info(f'[DIAG] _detectar_mejor_hoja: {total_sheets} hoja(s): {sheet_names}')

    for sname in sheet_names:
        _t_sheet = time.perf_counter()
        ws = wb[sname]
        raw = list(ws.iter_rows(values_only=True))
        _t_read = _diag_log(f'_detectar_mejor_hoja: leer filas [{sname}]', _t_sheet, f'filas={len(raw)}')
        score = matcher.score_sheet(sname, raw)
        _diag_log(f'_detectar_mejor_hoja: score_sheet [{sname}]', _t_read, f'score={score:.4f}')
        all_sheets_scores[sname] = score
        if score > best_score:
            best_score = score
            best_sheet_name = sname
            best_ws = ws

    if best_ws is None:
        raise ValueError(
            f'No se pudo detectar una hoja con datos válidos '
            f'entre las {total_sheets} hojas disponibles.'
        )

    logger.info(f'[DIAG] _detectar_mejor_hoja: mejor hoja="{best_sheet_name}" score={best_score:.4f} | total: {(time.perf_counter() - _t0) * 1000:.1f}ms')
    return best_ws, best_sheet_name, all_sheets_scores, total_sheets


def _analizar_encabezados_y_datos(
    ws, matcher: ColumnMatcher
) -> tuple[int, list[str], float, int, list[dict]]:
    """Detecta fila de encabezados, fila de datos y parsea filas."""
    _t0 = time.perf_counter()
    logger.info('[DIAG] >>> _analizar_encabezados_y_datos')

    _t_read = _diag_start('_analizar: list(iter_rows)')
    raw_all = list(ws.iter_rows(values_only=True))
    _diag_log('_analizar: list(iter_rows)', _t_read, f'total_filas={len(raw_all)}')

    _t_dbh = _diag_start('_analizar: detect_best_header_row')
    row_idx, headers, header_score = matcher.detect_best_header_row(raw_all)
    _diag_log('_analizar: detect_best_header_row', _t_dbh, f'row_idx={row_idx}, header_score={header_score:.4f}, headers={headers[:5]}...')

    if not any(h.strip() for h in headers if h):
        logger.warning('[DIAG] _analizar: No valid headers detected, falling back to row 0')
        row_idx = 0
        headers = (
            [str(c).strip() if c is not None else '' for c in raw_all[0]]
            if raw_all else []
        )
        header_score = 0.0

    headers = [h if h else '' for h in headers]

    _t_dsr = _diag_start('_analizar: detect_data_start_row')
    data_start_row = matcher.detect_data_start_row(raw_all, row_idx, headers)
    _diag_log('_analizar: detect_data_start_row', _t_dsr, f'data_start_row={data_start_row}')

    _t_pdr = _diag_start('_analizar: parse_data_rows')
    filas = parse_data_rows(ws, row_idx, headers, data_start_row)
    _diag_log('_analizar: parse_data_rows', _t_pdr, f'filas_parseadas={len(filas)}')

    if not filas:
        logger.warning(f'[DIAG] _analizar: No data rows from row {data_start_row}, trying fallback (row 1)')
        headers = (
            [str(c).strip() if c is not None else '' for c in raw_all[0]]
            if raw_all else []
        )
        data_start_row = 1
        filas = parse_data_rows(ws, 0, headers, 1)
        if not filas:
            raise ValueError(
                'No se encontraron filas de datos después de la fila de encabezados.'
            )
        logger.info(f'[DIAG] _analizar: fallback OK, filas={len(filas)}')

    logger.info(f'[DIAG] <<< _analizar_encabezados_y_datos | total: {(time.perf_counter() - _t0) * 1000:.1f}ms | filas={len(filas)}')
    return row_idx, headers, header_score, data_start_row, filas


def analyze_workbook(archivo, formulario) -> dict:
    """
    Analiza un archivo Excel con detección inteligente completa.

    Realiza:
      1. Detección automática de hoja.
      2. Detección automática de fila de encabezados.
      3. Detección de primera fila de datos.
      4. Matching inteligente de columnas 4 niveles (FASE 4).
      5. Cálculo de confianza por columna y global.
      6. Calidad del archivo (FASE 7).

    Returns:
        dict con análisis completo.
    """
    _t0 = time.perf_counter()
    logger.info('[DIAG] ========== analyze_workbook ENTRY ==========')

    _t_lw = _diag_start('analyze: load_workbook')
    try:
        wb = load_workbook(archivo, read_only=True, data_only=True)
    except Exception as e:
        _diag_log('analyze: load_workbook (FAILED)', _t_lw, str(e))
        raise ValueError(f'No se pudo leer el archivo Excel: {e}')
    _diag_log('analyze: load_workbook', _t_lw, f'sheets={wb.sheetnames}')

    _t_cm = _diag_start('analyze: ColumnMatcher(formulario)')
    matcher = ColumnMatcher(formulario)
    _diag_log('analyze: ColumnMatcher(formulario)', _t_cm, f'field_names={matcher.field_names}')

    try:
        _t_dh = _diag_start('analyze: _detectar_mejor_hoja')
        best_ws, best_sheet_name, all_sheets_scores, total_sheets = _detectar_mejor_hoja(wb, matcher)
        _diag_log('analyze: _detectar_mejor_hoja', _t_dh, f'best="{best_sheet_name}", scores={all_sheets_scores}')

        _t_ah = _diag_start('analyze: _analizar_encabezados_y_datos')
        row_idx, headers, header_score, data_start_row, filas = _analizar_encabezados_y_datos(best_ws, matcher)
        _diag_log('analyze: _analizar_encabezados_y_datos', _t_ah, f'headers={len(headers)}, filas={len(filas)}')

        _t_ma = _diag_start('analyze: match_all')
        match_results = matcher.match_all(headers)
        _diag_log('analyze: match_all', _t_ma, f'results={len(match_results)}')

        matched_cols = [r for r in match_results if r.matched_to]
        num_matched = len(matched_cols)
        num_fuzzy = sum(1 for r in matched_cols if r.method == 'fuzzy')
        num_synonym = sum(1 for r in matched_cols if r.method == 'synonym')
        num_exact = sum(1 for r in matched_cols if r.method == 'exact')
        num_normalized = sum(1 for r in matched_cols if r.method == 'normalized')
        logger.info(
            f'[DIAG] analyze: match_all summary — matched={num_matched}/{len(headers)}, '
            f'exact={num_exact}, normalized={num_normalized}, synonym={num_synonym}, fuzzy={num_fuzzy}'
        )

        confianza_global = (
            sum(r.confidence for r in matched_cols) / num_matched
            if matched_cols else 0.0
        )

        _t_cal = _diag_start('analyze: calcular_calidad')
        calidad = matcher.calcular_calidad(match_results, len(headers))
        _diag_log('analyze: calcular_calidad', _t_cal, f'estrellas={calidad.get("estrellas","?")}, score={calidad.get("score","?")}')

        conflictos_globales = [
            {'columna': headers[r.column_index], 'conflictos': r.conflicts}
            for r in match_results if r.conflicts
        ]

    finally:
        _t_close = _diag_start('analyze: wb.close')
        wb.close()
        _diag_log('analyze: wb.close', _t_close)

    logger.info(f'[DIAG] <<< analyze_workbook total: {(time.perf_counter() - _t0) * 1000:.1f}ms')
    logger.info(f'[DIAG] ========== analyze_workbook EXIT ==========')

    return {
        'sheet_name': best_sheet_name,
        'total_sheets': total_sheets,
        'all_sheets_scores': all_sheets_scores,
        'header_row': row_idx,
        'header_score': header_score,
        'data_start_row': data_start_row,
        'encabezados': headers,
        'filas': filas,
        'match_results': match_results,
        'total_filas': len(filas),
        'confianza_global': confianza_global * 100,
        'calidad': calidad,
        'conflictos_globales': conflictos_globales,
    }


# ======================================================================
# AUTO MAPPING — Orquestación completa: ColumnMatcher + MappingMemory +
#                 AIMatcher + AutoMappingAnalyzer
# ======================================================================


def analizar_y_clasificar_columnas(
    formulario,
    encabezados: list[str],
    campos_activos,
) -> dict:
    """
    Orquestación completa del análisis de mapeo automático.

    Flujo:
      1. AutoMappingAnalyzer.analyze_full():
         a. ColumnMatcher.match_all() — 4 niveles de matching clásico.
         b. MappingMemoryManager.load() — recupera mapeos previos.
         c. MappingMemoryManager.apply_memory_to_results() — aplica memoria.
         d. AIMatcher.match_unresolved() — matching semántico DeepSeek.
         e. AutoMappingAnalyzer.analyze() — clasificación final.
      2. Construcción del mapping final.
      3. Cómputo de campos sin mapear.

    Returns:
        dict con:
          - summary: MappingSummary con clasificación detallada.
          - mapeo_idx: dict {col_idx: campo_nombre} con el mapping final.
          - sin_mapear: list[str] de campos sin correspondencia.
    """
    _t0 = time.perf_counter()
    logger.info('[AUTO] ===== analizar_y_clasificar_columnas ENTRY =====')
    logger.info(f'[AUTO] formulario={formulario.nombre}, encabezados={len(encabezados)}, campos={campos_activos.count()}')

    from .auto_mapping import AutoMappingAnalyzer

    analyzer = AutoMappingAnalyzer()

    # Ejecutar orquestación completa
    summary = analyzer.analyze_full(
        formulario=formulario,
        encabezados=encabezados,
        campos_activos=campos_activos,
        excluded_types=TIPOS_EXCLUIDOS_MAPEO,
    )

    # Construir mapping final
    mapeo_idx = analyzer.build_mapping_from_summary(summary)

    # Calcular campos sin mapear (excluyendo tipos no mapeables)
    field_names = [c.nombre for c in campos_activos]
    sin_mapear = analyzer.get_fields_sin_mapear(
        summary, field_names, excluded_types=TIPOS_EXCLUIDOS_MAPEO
    )

    elapsed = (time.perf_counter() - _t0) * 1000
    logger.info(
        f'[AUTO] ===== analizar_y_clasificar_columnas EXIT: '
        f'{elapsed:.1f}ms '
        f'Auto={summary.auto}, Review={summary.review}, '
        f'Manual={summary.manual}, SinMapear={len(sin_mapear)}'
    )

    return {
        'summary': summary,
        'mapeo_idx': mapeo_idx,
        'sin_mapear': sin_mapear,
    }


def guardar_memoria_mapeo(
    formulario,
    encabezados: list[str],
    mapeo_idx: dict[int, str],
) -> bool:
    """
    Guarda el mapeo confirmado en la memoria persistente
    para reutilizarlo en futuras importaciones del mismo formulario.

    Args:
        formulario: Instancia de Formulario.
        encabezados: Lista de nombres de columna del Excel.
        mapeo_idx: Dict {col_idx: campo_nombre} del mapeo final.

    Returns:
        True si se guardó correctamente.
    """
    try:
        from .mapping_memory import MappingMemoryManager
        return MappingMemoryManager.save(
            formulario_id=formulario.id,
            encabezados=encabezados,
            mapping=mapeo_idx,
        )
    except Exception as e:
        logger.warning(f'Error guardando memoria de mapeo: {e}')
        return False


# ======================================================================
# Pipeline integration — future: fully replace `importar()` with this
# ======================================================================


def importar_con_pipeline(
    formulario, filepath, filename, usuario=None,
    modo='crear', mapping_override=None,
    sheet_name=None, header_row=None,
) -> dict:
    """Importa usando el pipeline empresarial (con auditoría y rollback)."""
    from .import_export.pipeline import ImportPipeline, PipelineConfig

    config = PipelineConfig(
        formulario_id=formulario.id,
        filepath=filepath,
        filename=filename,
        modo=modo,
        usuario_id=usuario.id if usuario else None,
        mapping_override=mapping_override,
        sheet_name=sheet_name,
        header_row=header_row,
    )
    pipeline = ImportPipeline()
    result = pipeline.run(config)

    return {
        'total': result.total_filas,
        'creados': result.creados,
        'actualizados': result.actualizados,
        'ignorados': result.ignorados,
        'errores': result.errors_detail,
        'tiempo_seg': result.tiempo_seg,
        'modo': modo,
        'import_log_id': result.import_log_id,
        'quality': result.quality,
        'warnings': result.warnings,
    }
