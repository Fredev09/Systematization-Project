import re

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ---------------------------------------------------------------------------
# Evaluación de fórmulas para campos calculados
# (movido aquí desde views.py para evitar dependencia service → view)
# ---------------------------------------------------------------------------


def _evaluar_formula(formula, valores_por_nombre):
    """Evalúa una fórmula matemática usando valores de campos del mismo formulario.
    Soporta: +, -, *, /, (, )
    Ejemplos: cantidad * precio_unitario, subtotal - descuento
    Los valores se pasan como dict {nombre_campo: valor_string}.
    """
    if not formula:
        return ''

    expr = formula.strip()

    # Reemplazar nombres de campo por sus valores numéricos
    for nombre, valor in valores_por_nombre.items():
        try:
            num_val = float(str(valor).replace(',', '.'))
        except (ValueError, TypeError):
            num_val = 0
        # Reemplazar palabra completa (no parcial)
        expr = re.sub(r'(?<![\w\d])' + re.escape(nombre) + r'(?![\w\d])',
                      str(num_val), expr)

    # Validar que solo queden números, operadores y paréntesis
    expr_limpia = expr.replace(' ', '')
    if not re.match(r'^[\d\+\-\*\/\(\)\.]+$', expr_limpia):
        return 'Error: fórmula inválida'

    try:
        resultado = eval(expr_limpia, {'__builtins__': {}}, {})
        if isinstance(resultado, (int, float)):
            redondeado = round(resultado, 2)
            # Si es un número entero, mostrarlo sin decimales
            if isinstance(redondeado, float) and redondeado == int(redondeado):
                return str(int(redondeado))
            return str(redondeado)
        return str(resultado)
    except Exception:
        return 'Error al calcular'


def exportar_registros_excel(registros, campos, formulario_nombre, relacion_resolver=None):
    """
    Genera un archivo Excel con los registros de un formulario dinámico.

    Args:
        registros: QuerySet de Registro
        campos: QuerySet de Campo ordenados
        formulario_nombre: Nombre del formulario para el título
        relacion_resolver: Función opcional para resolver campos tipo 'relacion'
                          (campos, valores_map) -> {registro_id: display_text}

    Returns:
        HttpResponse con el archivo Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Registros'

    # Identificar campo identificador principal
    campo_identificador = campos.filter(identificador_principal=True).first()

    # Estilos
    titulo_fill = PatternFill('solid', fgColor='D41473')
    encabezado_fill = PatternFill('solid', fgColor='FCE7F3')
    resumen_fill = PatternFill('solid', fgColor='FFF1F8')
    borde_color = Side(style='thin', color='E5E7EB')

    titulo_font = Font(color='FFFFFF', bold=True, size=15)
    encabezado_font = Font(color='111827', bold=True)
    resumen_font = Font(color='4B5563', bold=True)
    texto_font = Font(color='111827')

    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')

    # Construir encabezados (sin ID interno)
    encabezados = ['Fecha']
    if campo_identificador:
        encabezados.append(campo_identificador.nombre)
    for campo in campos:
        if campo.id != getattr(campo_identificador, 'id', None):
            encabezados.append(campo.nombre)
    encabezados.append('Usuario')

    # Título y resumen
    ws.append([f'Registros: {formulario_nombre}'])
    ws.append([
        f'Total de registros: {registros.count()}',
        f'Generado: {timezone.localtime(timezone.now()).strftime("%d/%m/%Y %I:%M %p")}'
    ])
    ws.append([])
    ws.append(encabezados)

    # Mapa de valores por registro: {registro_id: {campo_id: valor}}
    from .models import ValorCampo
    valores_qs = ValorCampo.objects.filter(
        registro__in=registros
    ).select_related('campo', 'registro')

    valores_map = {}
    for vc in valores_qs:
        if vc.registro_id not in valores_map:
            valores_map[vc.registro_id] = {}
        valores_map[vc.registro_id][vc.campo_id] = vc.valor

    # Resolver campos tipo 'relacion' si hay resolver
    if relacion_resolver:
        relacion_resuelto = relacion_resolver(campos, valores_map)
        for reg_id, vals in valores_map.items():
            for campo in campos:
                if campo.tipo == 'relacion':
                    raw = vals.get(campo.id, '')
                    if raw and raw.isdigit():
                        ref_id = int(raw)
                        display = relacion_resuelto.get(ref_id)
                        if display:
                            vals[campo.id] = display

    # Datos
    campos_ids = [c.id for c in campos]
    id_identificador = getattr(campo_identificador, 'id', None) if campo_identificador else None
    for registro in registros:
        fecha = timezone.localtime(registro.fecha_creacion).strftime(
            '%d/%m/%Y %I:%M %p'
        )
        usuario = registro.usuario.username if registro.usuario else ''

        fila = [fecha]
        valores_reg = valores_map.get(registro.id, {})
        if id_identificador:
            fila.append(valores_reg.get(id_identificador, ''))
        for campo_id in campos_ids:
            if campo_id != id_identificador:
                fila.append(valores_reg.get(campo_id, ''))

        fila.append(usuario)
        ws.append(fila)

    # Formato del título
    num_columnas = len(encabezados)
    ws.merge_cells(
        start_row=1, start_column=1,
        end_row=1, end_column=num_columnas
    )
    ws['A1'].fill = titulo_fill
    ws['A1'].font = titulo_font
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 28

    # Formato del resumen
    for cell in ws[2]:
        cell.fill = resumen_fill
        cell.font = resumen_font
        cell.alignment = left
        cell.border = Border(
            left=borde_color, right=borde_color,
            top=borde_color, bottom=borde_color
        )

    # Formato de encabezados
    for cell in ws[4]:
        cell.fill = encabezado_fill
        cell.font = encabezado_font
        cell.alignment = center
        cell.border = Border(
            left=borde_color, right=borde_color,
            top=borde_color, bottom=borde_color
        )

    # Formato de datos
    for row in ws.iter_rows(min_row=5):
        for cell in row:
            cell.font = texto_font
            cell.alignment = left
            cell.border = Border(
                left=borde_color, right=borde_color,
                top=borde_color, bottom=borde_color
            )

    def _col_letter(n):
        letters = ''
        while n > 0:
            n -= 1
            letters = chr(65 + (n % 26)) + letters
            n //= 26
        return letters

    # Anchos de columna
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 22
    for i, _ in enumerate(campos, start=3):
        col_letter = _col_letter(i)
        ws.column_dimensions[col_letter].width = 28

    ultima_col = _col_letter(num_columnas)

    if num_columnas >= 3:
        letra_usuario = _col_letter(num_columnas)
        ws.column_dimensions[letra_usuario].width = 22

    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:{ultima_col}{ws.max_row}'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="registros_{formulario_nombre.lower().replace(" ", "_")}.xlsx"'
    )

    wb.save(response)
    return response
