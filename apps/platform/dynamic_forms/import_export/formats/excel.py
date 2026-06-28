"""Excel parser using openpyxl."""

from __future__ import annotations

from pathlib import Path
from datetime import datetime, date
from typing import Any

from openpyxl import load_workbook

from .base import BaseParser, ParseResult


class ExcelParser(BaseParser):
    """Excel (.xlsx) parser with sheet/row detection."""

    NOISE_PATTERNS: set[str] = {
        'empresa', 'reporte', 'resumen', 'totales', 'total',
        'company', 'summary', 'report', 'total', 'subtotal',
        'generado', 'fecha', 'date', 'página', 'page',
        'instrucciones', 'instructions', 'notas', 'notes',
    }

    def _load(self, data_only: bool = True, read_only: bool = True):
        return load_workbook(
            self.filepath,
            data_only=data_only,
            read_only=read_only,
        )

    def _es_fila_vacia(self, row: tuple) -> bool:
        return all(c.value is None or str(c.value).strip() == '' for c in row)

    def _es_fila_separacion(self, row: tuple) -> bool:
        vals = [str(c.value or '').strip() for c in row if c.value is not None]
        return bool(vals) and all(v in ('---', '___', '...', '***', '') for v in vals)

    def _es_fila_titulo_repetido(self, row: tuple, headers: list[str]) -> bool:
        vals = [str(c.value or '').strip() for c in row]
        return len(vals) > 0 and vals == headers

    def _es_fila_sumario(self, row: tuple) -> bool:
        first = str(row[0].value or '').strip().lower() if row else ''
        return any(k in first for k in ('total', 'subtotal', 'sum', 'promedio', 'average'))

    def _es_fila_ruido(self, row: tuple) -> bool:
        first = str(row[0].value or '').strip().lower() if row else ''
        return any(k in first for k in self.NOISE_PATTERNS)

    def _valor_celda(self, cell) -> str:
        val = cell.value
        if val is None:
            return ''
        if isinstance(val, datetime):
            return val.strftime('%Y-%m-%d %H:%M:%S')
        if isinstance(val, date):
            return val.isoformat()
        return str(val)

    def score_sheet(self, sheet, field_names: list[str]) -> float:
        score = 0.0
        for r_idx, row in enumerate(sheet.iter_rows(max_row=20, values_only=False)):
            if r_idx > 15:
                break
            vals = [str(c.value or '').strip().lower() for c in row if c.value is not None]
            for fn in field_names:
                fn_lower = fn.lower()
                if fn_lower in vals:
                    score += 2.0
        sheet_name_lower = sheet.title.lower()
        if any(k in sheet_name_lower for k in ('resumen', 'summary', 'instrucciones', 'instructions', 'notas')):
            score -= 5.0
        if any(k in sheet_name_lower for k in ('datos', 'data', 'ventas', 'productos', 'clientes', 'hoja1', 'sheet1')):
            score += 3.0
        return max(score, 0.0)

    def detect_best_header_row(self, sheet, field_names: list[str]) -> tuple[int, float]:
        best_row = 0
        best_score = 0.0
        fn_lower = [f.lower() for f in field_names]
        for r_idx, row in enumerate(sheet.iter_rows(max_row=20, values_only=False)):
            if r_idx > 18:
                break
            vals = [str(c.value or '').strip().lower() for c in row if c.value is not None]
            if not vals:
                continue
            if self._es_fila_ruido(row):
                continue
            score = sum(2.0 for v in vals if v in fn_lower)
            if score > best_score:
                best_score = score
                best_row = r_idx
        return best_row, best_score

    def detect_data_start_row(self, sheet, header_row: int) -> int:
        start = header_row + 1
        for r_idx, row in enumerate(sheet.iter_rows(min_row=start, max_row=start + 10, values_only=False), start=start):
            if self._es_fila_vacia(row):
                continue
            if self._es_fila_separacion(row):
                continue
            if self._es_fila_sumario(row):
                continue
            return r_idx
        return start

    def detect_structure(self) -> dict:
        wb = self._load()
        info = {
            'total_sheets': len(wb.sheetnames),
            'sheets': [],
        }
        for name in wb.sheetnames:
            ws = wb[name]
            info['sheets'].append({
                'name': name,
                'rows': ws.max_row or 0,
                'cols': ws.max_column or 0,
            })
        wb.close()
        return info

    def parse(
        self,
        sheet_name: str | None = None,
        header_row: int | None = None,
        field_names: list[str] | None = None,
    ) -> ParseResult:
        wb = self._load()

        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        field_names = field_names or []
        hdr_row = header_row if header_row is not None else self.detect_best_header_row(ws, field_names)[0]
        data_start = self.detect_data_start_row(ws, hdr_row)

        rows_iter = ws.iter_rows(min_row=hdr_row + 1, values_only=False)
        header_cells = list(ws.iter_rows(min_row=hdr_row + 1, max_row=hdr_row + 1, values_only=False))[0]

        headers = [str(c.value or '').strip() if c.value is not None else '' for c in header_cells]
        rows = []
        for row in ws.iter_rows(min_row=data_start, values_only=False):
            if self._es_fila_vacia(row):
                continue
            if self._es_fila_separacion(row):
                continue
            if self._es_fila_sumario(row):
                continue
            rows.append([self._valor_celda(c) for c in row])

        wb.close()

        return ParseResult(
            headers=headers,
            rows=rows,
            sheet_name=ws.title,
            header_row=hdr_row,
            data_start_row=data_start,
            total_sheets=len(wb.sheetnames),
            filename=self.filename,
        )
